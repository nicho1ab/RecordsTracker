"""Focused unit tests for the stakeholder facility overview extract module."""
from __future__ import annotations

import csv
import json
import sqlite3
from pathlib import Path

import openpyxl

from ccld_complaints.stakeholder_extract import (
    UNKNOWN,
    FacilityReferenceError,
    FacilityReferenceFilterError,
    export_stakeholder_facility_overview,
    is_substantiated_equivalent,
    read_facility_reference_csv,
)

# ---------------------------------------------------------------------------
# Helper: in-memory SQLite with minimal schema
# ---------------------------------------------------------------------------

_MINI_SCHEMA = """
CREATE TABLE IF NOT EXISTS facilities (
    facility_id TEXT PRIMARY KEY,
    source_id TEXT NOT NULL,
    external_facility_number TEXT NOT NULL,
    facility_name TEXT,
    facility_type TEXT,
    licensee_name TEXT,
    county TEXT,
    status TEXT,
    capacity INTEGER,
    regional_office TEXT
);

CREATE TABLE IF NOT EXISTS source_documents (
    document_id TEXT PRIMARY KEY,
    source_id TEXT NOT NULL,
    facility_id TEXT NOT NULL,
    source_url TEXT NOT NULL UNIQUE,
    retrieved_at TEXT NOT NULL,
    raw_sha256 TEXT,
    connector_name TEXT NOT NULL,
    connector_version TEXT NOT NULL,
    raw_path TEXT,
    document_type TEXT,
    report_index INTEGER,
    http_status INTEGER,
    content_type TEXT
);

CREATE TABLE IF NOT EXISTS complaints (
    complaint_id TEXT PRIMARY KEY,
    facility_id TEXT NOT NULL,
    document_id TEXT NOT NULL,
    complaint_control_number TEXT,
    complaint_received_date TEXT,
    first_investigation_activity_date TEXT,
    visit_date TEXT,
    report_date TEXT,
    date_signed TEXT,
    finding TEXT,
    days_received_to_first_activity INTEGER,
    days_received_to_visit INTEGER,
    days_received_to_report INTEGER,
    days_report_to_signed INTEGER,
    review_delay_over_30_days INTEGER NOT NULL DEFAULT 0,
    review_delay_over_60_days INTEGER NOT NULL DEFAULT 0,
    review_delay_over_90_days INTEGER NOT NULL DEFAULT 0,
    review_delay_over_120_days INTEGER NOT NULL DEFAULT 0,
    missing_first_activity_date INTEGER NOT NULL DEFAULT 0,
    report_date_used_as_proxy INTEGER NOT NULL DEFAULT 0,
    extraction_confidence REAL
);

CREATE TABLE IF NOT EXISTS allegations (
    allegation_id TEXT PRIMARY KEY,
    complaint_id TEXT NOT NULL,
    allegation_text TEXT NOT NULL,
    allegation_category TEXT,
    finding TEXT,
    extraction_confidence REAL
);
"""


def _make_db() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.executescript(_MINI_SCHEMA)
    conn.row_factory = sqlite3.Row
    return conn


def _insert_facility(
    conn: sqlite3.Connection, *, fid: str, fnum: str, name: str | None = "Test Facility"
) -> None:
    conn.execute(
        """
        INSERT INTO facilities
            (facility_id, source_id, external_facility_number, facility_name,
             facility_type, status, county)
        VALUES (?, 'ccld', ?, ?, 'Child Care Center', 'Licensed', 'Alameda')
        """,
        (fid, fnum, name),
    )


def _insert_source_doc(
    conn: sqlite3.Connection, *, doc_id: str, fid: str, url: str, sha256: str = "abc123"
) -> None:
    conn.execute(
        """
        INSERT INTO source_documents
            (document_id, source_id, facility_id, source_url, retrieved_at,
             raw_sha256, connector_name, connector_version)
        VALUES (?, 'ccld', ?, ?, '2024-01-01T00:00:00Z', ?, 'ccld', '1.0')
        """,
        (doc_id, fid, url, sha256),
    )


def _insert_complaint(
    conn: sqlite3.Connection,
    *,
    cid: str,
    fid: str,
    doc_id: str,
    control_number: str,
    finding: str = "Unsubstantiated",
    received_date: str = "2022-04-07",
    report_date: str = "2022-06-01",
) -> None:
    conn.execute(
        """
        INSERT INTO complaints
            (complaint_id, facility_id, document_id, complaint_control_number,
             complaint_received_date, report_date, finding)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (cid, fid, doc_id, control_number, received_date, report_date, finding),
    )


# ---------------------------------------------------------------------------
# is_substantiated_equivalent
# ---------------------------------------------------------------------------

class TestIsSubstantiatedEquivalent:
    def test_substantiated_matches(self) -> None:
        assert is_substantiated_equivalent("Substantiated") is True

    def test_founded_matches(self) -> None:
        assert is_substantiated_equivalent("Founded") is True

    def test_sustained_matches(self) -> None:
        assert is_substantiated_equivalent("Sustained") is True

    def test_case_insensitive(self) -> None:
        assert is_substantiated_equivalent("SUBSTANTIATED") is True

    def test_unsubstantiated_excluded(self) -> None:
        assert is_substantiated_equivalent("Unsubstantiated") is False

    def test_not_substantiated_excluded(self) -> None:
        assert is_substantiated_equivalent("Not Substantiated") is False

    def test_empty_string_false(self) -> None:
        assert is_substantiated_equivalent("") is False

    def test_none_false(self) -> None:
        assert is_substantiated_equivalent(None) is False

    def test_unrelated_string_false(self) -> None:
        assert is_substantiated_equivalent("Pending") is False


# ---------------------------------------------------------------------------
# Test 1: Facility aggregation counts loaded complaints correctly
# ---------------------------------------------------------------------------

class TestFacilityAggregation:
    def test_counts_loaded_complaints_correctly(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_facility(conn, fid="f2", fnum="100002")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_source_doc(conn, doc_id="d2", fid="f1", url="http://example.test/2")
        _insert_source_doc(conn, doc_id="d3", fid="f2", url="http://example.test/3")
        _insert_complaint(conn, cid="c1", fid="f1", doc_id="d1", control_number="CR-001")
        _insert_complaint(conn, cid="c2", fid="f1", doc_id="d2", control_number="CR-002")
        _insert_complaint(conn, cid="c3", fid="f2", doc_id="d3", control_number="CR-003")
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        # Materialise to file so export_stakeholder_facility_overview can open it
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        assert result.facility_row_count == 2
        # Read CSV to verify counts
        rows = _read_csv(result.facility_overview_path)
        by_num = {r["FacilityNumber"]: r for r in rows}
        assert by_num["100001"]["LoadedComplaintCount"] == "2"
        assert by_num["100002"]["LoadedComplaintCount"] == "1"

    def test_substantiated_count_correct(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_source_doc(conn, doc_id="d2", fid="f1", url="http://example.test/2")
        _insert_source_doc(conn, doc_id="d3", fid="f1", url="http://example.test/3")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        _insert_complaint(
            conn, cid="c2", fid="f1", doc_id="d2",
            control_number="CR-002", finding="Unsubstantiated",
        )
        _insert_complaint(
            conn, cid="c3", fid="f1", doc_id="d3",
            control_number="CR-003", finding="Founded",
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        rows = _read_csv(result.facility_overview_path)
        assert rows[0]["SubstantiatedOrEquivalentCount"] == "2"


# ---------------------------------------------------------------------------
# Test 2: Substantiated/equivalent complaints are counted and listed
# ---------------------------------------------------------------------------

class TestSubstantiatedComplaints:
    def test_substantiated_listed_in_substantiated_csv(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001", name="Alpha Care Center")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/report1")
        _insert_source_doc(conn, doc_id="d2", fid="f1", url="http://example.test/report2")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-SUB-001", finding="Substantiated",
        )
        _insert_complaint(
            conn, cid="c2", fid="f1", doc_id="d2",
            control_number="CR-UNSUB-001", finding="Unsubstantiated",
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        assert result.substantiated_complaint_row_count == 1
        rows = _read_csv(result.substantiated_complaints_path)
        assert len(rows) == 1
        assert rows[0]["ComplaintControlNumber"] == "CR-SUB-001"
        assert rows[0]["FindingOrResolution"] == "Substantiated"
        assert rows[0]["FacilityName"] == "Alpha Care Center"

    def test_founded_and_sustained_also_listed(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_source_doc(conn, doc_id="d2", fid="f1", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Founded",
        )
        _insert_complaint(
            conn, cid="c2", fid="f1", doc_id="d2",
            control_number="CR-002", finding="Sustained",
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        assert result.substantiated_complaint_row_count == 2


# ---------------------------------------------------------------------------
# Test 3: Missing facility name/date/source URL uses safe fallback text
# ---------------------------------------------------------------------------

class TestSafeFallbacks:
    def test_null_facility_name_uses_unknown(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001", name=None)
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        facility_rows = _read_csv(result.facility_overview_path)
        sub_rows = _read_csv(result.substantiated_complaints_path)
        assert facility_rows[0]["FacilityName"] == UNKNOWN
        assert sub_rows[0]["FacilityName"] == UNKNOWN

    def test_null_dates_use_unknown(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        conn.execute(
            """
            INSERT INTO complaints
                (complaint_id, facility_id, document_id, complaint_control_number, finding)
            VALUES ('c1', 'f1', 'd1', 'CR-001', 'Substantiated')
            """
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        facility_rows = _read_csv(result.facility_overview_path)
        sub_rows = _read_csv(result.substantiated_complaints_path)
        assert facility_rows[0]["EarliestComplaintDate"] == UNKNOWN
        assert facility_rows[0]["MostRecentComplaintDate"] == UNKNOWN
        assert sub_rows[0]["ComplaintReceivedDate"] == UNKNOWN
        assert sub_rows[0]["ReportDate"] == UNKNOWN

    def test_null_source_url_uses_unknown(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        conn.execute(
            """
            INSERT INTO source_documents
                (document_id, source_id, facility_id, source_url, retrieved_at,
                 raw_sha256, connector_name, connector_version)
            VALUES ('d1', 'ccld', 'f1', 'http://example.test/1', '2024-01-01T00:00:00Z',
                    NULL, 'ccld', '1.0')
            """
        )
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        sub_rows = _read_csv(result.substantiated_complaints_path)
        assert sub_rows[0]["RawHashOrArtifactReference"] == UNKNOWN


# ---------------------------------------------------------------------------
# Test 4: Empty input produces headers, README, manifest, and ZIP
# ---------------------------------------------------------------------------

class TestEmptyInput:
    def test_nonexistent_db_produces_valid_empty_outputs(self, tmp_path: Path) -> None:
        db_path = tmp_path / "does_not_exist.sqlite"
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        assert result.facility_row_count == 0
        assert result.substantiated_complaint_row_count == 0

        # CSV files exist and have correct headers
        facility_rows = _read_csv_raw(result.facility_overview_path)
        assert facility_rows[0] == list(_facility_overview_headers())
        assert len(facility_rows) == 1  # header only

        sub_rows = _read_csv_raw(result.substantiated_complaints_path)
        assert sub_rows[0] == list(_substantiated_complaints_headers())
        assert len(sub_rows) == 1  # header only

    def test_empty_output_readme_has_required_caution_language(self, tmp_path: Path) -> None:
        db_path = tmp_path / "does_not_exist.sqlite"
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        readme = result.readme_path.read_text(encoding="utf-8")
        assert "source of record" in readme
        assert "does not make legal conclusions" in readme
        assert "does not make facility-wide conclusions" in readme
        assert "not independently verified" in readme
        assert "raw narrative" in readme
        assert "does not claim to be source-complete" in readme
        assert "zero" in readme.lower() or "empty" in readme.lower() or "counts" in readme.lower()

    def test_empty_output_manifest_counts_are_zero(self, tmp_path: Path) -> None:
        db_path = tmp_path / "does_not_exist.sqlite"
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        assert manifest["facility_row_count"] == 0
        assert manifest["substantiated_complaint_row_count"] == 0
        assert "limitations" in manifest
        assert manifest["script_name"] == "export-stakeholder-facility-overview.ps1"

    def test_empty_output_xlsx_created(self, tmp_path: Path) -> None:
        db_path = tmp_path / "does_not_exist.sqlite"
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        assert result.xlsx_path.exists()
        wb = openpyxl.load_workbook(result.xlsx_path)
        assert "README" in wb.sheetnames
        assert "Facility Overview" in wb.sheetnames
        assert "Substantiated Complaints" in wb.sheetnames
        assert "Manifest" in wb.sheetnames

    def test_empty_output_no_misleading_count_claims(self, tmp_path: Path) -> None:
        db_path = tmp_path / "does_not_exist.sqlite"
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        readme = result.readme_path.read_text(encoding="utf-8")
        # README must not claim zero means no complaints exist
        assert "do not prove" in readme or "does not prove" in readme or "not prove" in readme


# ---------------------------------------------------------------------------
# Test 5: Raw narrative fields not included in either CSV
# ---------------------------------------------------------------------------

class TestNoRawNarrative:
    def test_allegation_text_not_in_facility_overview_csv(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(conn, cid="c1", fid="f1", doc_id="d1", control_number="CR-001")
        conn.execute(
            """
            INSERT INTO allegations (allegation_id, complaint_id, allegation_text)
            VALUES ('a1', 'c1', 'SENTINEL_NARRATIVE_TEXT_SHOULD_NOT_APPEAR')
            """
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        content = result.facility_overview_path.read_text(encoding="utf-8")
        assert "SENTINEL_NARRATIVE_TEXT_SHOULD_NOT_APPEAR" not in content

    def test_allegation_text_not_in_substantiated_csv(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        conn.execute(
            """
            INSERT INTO allegations (allegation_id, complaint_id, allegation_text)
            VALUES ('a1', 'c1', 'SENTINEL_NARRATIVE_TEXT_SHOULD_NOT_APPEAR')
            """
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        content = result.substantiated_complaints_path.read_text(encoding="utf-8")
        assert "SENTINEL_NARRATIVE_TEXT_SHOULD_NOT_APPEAR" not in content

    def test_csv_headers_contain_no_narrative_column(self, tmp_path: Path) -> None:
        db_path = tmp_path / "does_not_exist.sqlite"
        export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        facility_headers = _facility_overview_headers()
        sub_headers = _substantiated_complaints_headers()

        narrative_keywords = {"text", "narrative", "allegation_text", "description", "summary"}
        for header in facility_headers:
            assert header.lower() not in narrative_keywords, f"Unexpected column: {header}"
        for header in sub_headers:
            assert header.lower() not in narrative_keywords, f"Unexpected column: {header}"


# ---------------------------------------------------------------------------
# Test 6: Manifest counts match generated CSV row counts
# ---------------------------------------------------------------------------

class TestManifestCounts:
    def test_manifest_counts_match_csv_row_counts(self, tmp_path: Path) -> None:
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_facility(conn, fid="f2", fnum="100002")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_source_doc(conn, doc_id="d2", fid="f2", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        _insert_complaint(
            conn, cid="c2", fid="f2", doc_id="d2",
            control_number="CR-002", finding="Unsubstantiated",
        )
        conn.commit()

        db_path = tmp_path / "test.sqlite"
        file_conn = sqlite3.connect(db_path)
        conn.backup(file_conn)
        file_conn.close()

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))

        facility_csv_rows = len(_read_csv(result.facility_overview_path))
        sub_csv_rows = len(_read_csv(result.substantiated_complaints_path))

        assert manifest["facility_row_count"] == facility_csv_rows
        assert manifest["substantiated_complaint_row_count"] == sub_csv_rows
        assert manifest["facility_row_count"] == result.facility_row_count
        sub_count = result.substantiated_complaint_row_count
        assert manifest["substantiated_complaint_row_count"] == sub_count

    def test_manifest_output_files_list_matches_actual_files(self, tmp_path: Path) -> None:
        db_path = tmp_path / "does_not_exist.sqlite"
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        listed = set(manifest["output_files"])
        expected = {
            "facility-overview.csv",
            "substantiated-complaints.csv",
            "README.md",
            "manifest.json",
        }
        assert expected.issubset(listed)
        # At least one .xlsx file listed (ZIP replaced by XLSX deliverable)
        assert any(f.endswith(".xlsx") for f in listed)
        # No .zip file listed
        assert not any(f.endswith(".zip") for f in listed)


class TestLimitationsSentence:
    """Regression guard for the canonical limitations sentence text."""

    def test_exact_phrase_source_completeness_claims(self) -> None:
        """_limitations_sentence contains 'source-completeness claims' with the space."""
        from ccld_complaints.stakeholder_extract import _limitations_sentence

        lim = _limitations_sentence()
        assert "source-completeness claims" in lim, (
            f"Expected 'source-completeness claims' in limitations sentence; got: {lim!r}"
        )

    def test_no_typo_source_completenessclaims(self) -> None:
        """_limitations_sentence must not contain the run-together typo."""
        from ccld_complaints.stakeholder_extract import _limitations_sentence

        lim = _limitations_sentence()
        assert "source-completenessclaims" not in lim, (
            f"Typo 'source-completenessclaims' found in limitations sentence: {lim!r}"
        )

    def test_generated_manifest_limitations_has_exact_phrase(self, tmp_path: Path) -> None:
        """Manifest limitations field in the generated export uses the correct phrase."""
        db_path = tmp_path / "does_not_exist.sqlite"
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        import json
        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        lim = manifest["limitations"]
        assert "source-completeness claims" in lim
        assert "source-completenessclaims" not in lim


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------


class TestFacilityReferenceInput:
    def test_reference_only_facility_has_zero_loaded_counts(
        self, tmp_path: Path
    ) -> None:
        """A reference-only facility (no loaded complaints) appears with zero counts."""
        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "999001", "FacilityName": "Reference Only Center"}],
            header=["FacilityNumber", "FacilityName"],
        )
        db_path = tmp_path / "does_not_exist.sqlite"

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
        )

        rows = _read_csv(result.facility_overview_path)
        assert len(rows) == 1
        row = rows[0]
        assert row["FacilityNumber"] == "999001"
        assert row["FacilityName"] == "Reference Only Center"
        assert row["LoadedComplaintCount"] == "0"
        assert row["SubstantiatedOrEquivalentCount"] == "0"
        assert row["ComplaintDataLoadedStatus"] == "No complaints loaded"

    def test_reference_only_facility_limitations_do_not_imply_no_complaints(
        self, tmp_path: Path
    ) -> None:
        """Limitations text must not imply zero means no public complaints exist."""
        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "999001", "FacilityName": "Test Center"}],
            header=["FacilityNumber", "FacilityName"],
        )
        db_path = tmp_path / "does_not_exist.sqlite"

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
        )

        rows = _read_csv(result.facility_overview_path)
        limitations = rows[0]["Limitations"]
        # Must carry the exact canonical limitations phrase (guards against typo like
        # "source-completenessclaims" that would collapse the space between words)
        assert "source-completeness claims" in limitations
        assert "source-completenessclaims" not in limitations

    def test_complaint_loaded_facility_merges_reference_metadata(
        self, tmp_path: Path
    ) -> None:
        """Reference metadata (city) enriches a facility that has loaded complaints."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001", name="Loaded Center")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{
                "FacilityNumber": "100001",
                "FacilityName": "Loaded Center",
                "City": "Sacramento",
                "County": "Sacramento County",
                "Status": "Licensed",
            }],
            header=["FacilityNumber", "FacilityName", "City", "County", "Status"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
        )

        rows = _read_csv(result.facility_overview_path)
        assert len(rows) == 1
        row = rows[0]
        assert row["FacilityNumber"] == "100001"
        assert row["LoadedComplaintCount"] == "1"
        assert row["SubstantiatedOrEquivalentCount"] == "1"
        # City is reference-only (not in DB schema); it should come from reference CSV
        assert row["City"] == "Sacramento"

    def test_missing_facility_number_column_raises_clearly(
        self, tmp_path: Path
    ) -> None:
        """CSV without a recognised facility number column raises FacilityReferenceError."""
        ref_csv = tmp_path / "bad.csv"
        _write_ref_csv(
            ref_csv,
            [{"Name": "Test", "Type": "Child Care"}],
            header=["Name", "Type"],
        )

        import pytest

        with pytest.raises(FacilityReferenceError) as exc_info:
            read_facility_reference_csv(ref_csv)

        assert "facility" in str(exc_info.value).casefold()
        assert "number" in str(exc_info.value).casefold()

    def test_duplicate_reference_rows_keep_first(self, tmp_path: Path) -> None:
        """When the reference CSV has duplicate facility numbers, first row wins."""
        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [
                {"FacilityNumber": "100001", "FacilityName": "First Name"},
                {"FacilityNumber": "100001", "FacilityName": "Second Name"},
            ],
            header=["FacilityNumber", "FacilityName"],
        )

        records = read_facility_reference_csv(ref_csv)

        assert len(records) == 1
        assert records[0]["facility_name"] == "First Name"

    def test_raw_ccld_headers_recognised(self, tmp_path: Path) -> None:
        """Raw CCLD column headers (FAC_NBR, NAME, etc.) are accepted as aliases."""
        ref_csv = tmp_path / "facility-reference.csv"
        _write_ref_csv(
            ref_csv,
            [{
                "FAC_NBR": "200001",
                "NAME": "Raw CCLD Center",
                "FAC_TYPE_DESC": "Child Care Center",
                "PROGRAM_TYPE": "Child Day Care",
                "STATUS": "Licensed",
                "RES_CITY": "Fresno",
                "COUNTY": "Fresno",
            }],
            header=[
                "FAC_NBR", "NAME", "FAC_TYPE_DESC", "PROGRAM_TYPE",
                "STATUS", "RES_CITY", "COUNTY",
            ],
        )

        records = read_facility_reference_csv(ref_csv)

        assert len(records) == 1
        rec = records[0]
        assert rec["facility_number"] == "200001"
        assert rec["facility_name"] == "Raw CCLD Center"
        assert rec["facility_type"] == "Child Care Center"
        assert rec["status"] == "Licensed"
        assert rec["city"] == "Fresno"
        assert rec["county"] == "Fresno"

    def test_raw_ccld_reference_only_facility_in_output(
        self, tmp_path: Path
    ) -> None:
        """A reference-only facility supplied via raw CCLD headers appears in output."""
        ref_csv = tmp_path / "facility-reference.csv"
        _write_ref_csv(
            ref_csv,
            [{"FAC_NBR": "300001", "NAME": "CCLD Only Facility", "RES_CITY": "Modesto"}],
            header=["FAC_NBR", "NAME", "RES_CITY"],
        )
        db_path = tmp_path / "does_not_exist.sqlite"

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
        )

        rows = _read_csv(result.facility_overview_path)
        assert len(rows) == 1
        row = rows[0]
        assert row["FacilityNumber"] == "300001"
        assert row["FacilityName"] == "CCLD Only Facility"
        assert row["City"] == "Modesto"
        assert row["LoadedComplaintCount"] == "0"

    def test_reference_input_no_raw_narrative_in_output(
        self, tmp_path: Path
    ) -> None:
        """Raw narrative text from allegations is not present even with reference input."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        conn.execute(
            """
            INSERT INTO allegations (allegation_id, complaint_id, allegation_text)
            VALUES ('a1', 'c1', 'SENTINEL_NARRATIVE_REFERENCE_TEST')
            """
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "100001", "FacilityName": "Test Center"}],
            header=["FacilityNumber", "FacilityName"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
        )

        for output_file in [
            result.facility_overview_path,
            result.substantiated_complaints_path,
        ]:
            assert "SENTINEL_NARRATIVE_REFERENCE_TEST" not in output_file.read_text(
                encoding="utf-8"
            )

    def test_manifest_includes_reference_counts(self, tmp_path: Path) -> None:
        """Manifest includes facility_reference_csv, _row_count, _matched_count."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [
                {"FacilityNumber": "100001", "FacilityName": "Matched"},
                {"FacilityNumber": "999999", "FacilityName": "Unmatched"},
            ],
            header=["FacilityNumber", "FacilityName"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
        )

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        assert "facility_reference_csv" in manifest
        assert manifest["facility_reference_row_count"] == 2
        assert manifest["facility_reference_matched_count"] == 1
        # Total facility rows = 1 loaded + 1 reference-only = 2
        assert result.facility_row_count == 2


# ---------------------------------------------------------------------------
# Tests: reference-only filter (--only-facility-reference-rows)
# ---------------------------------------------------------------------------


class TestOnlyFacilityReferenceRows:
    def test_unrelated_loaded_facility_excluded(self, tmp_path: Path) -> None:
        """Loaded facilities not in the reference CSV are excluded."""
        conn = _make_db()
        # Reference facility
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        # Unrelated loaded facility (not in reference)
        _insert_facility(conn, fid="f2", fnum="999999")
        _insert_source_doc(conn, doc_id="d2", fid="f2", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c2", fid="f2", doc_id="d2",
            control_number="CR-002", finding="Unsubstantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "100001", "FacilityName": "In Reference"}],
            header=["FacilityNumber", "FacilityName"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
            only_facility_reference_rows=True,
        )

        rows = _read_csv(result.facility_overview_path)
        facility_numbers = {r["FacilityNumber"] for r in rows}
        assert "100001" in facility_numbers
        assert "999999" not in facility_numbers
        assert result.facility_row_count == 1

    def test_reference_facility_with_loaded_complaints_included(
        self, tmp_path: Path
    ) -> None:
        """A reference facility that has loaded complaints is retained."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "100001", "FacilityName": "In Reference"}],
            header=["FacilityNumber", "FacilityName"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
            only_facility_reference_rows=True,
        )

        rows = _read_csv(result.facility_overview_path)
        assert len(rows) == 1
        assert rows[0]["FacilityNumber"] == "100001"
        assert rows[0]["LoadedComplaintCount"] == "1"

    def test_reference_facility_without_loaded_complaints_included(
        self, tmp_path: Path
    ) -> None:
        """A reference-only facility (zero loaded complaints) is retained."""
        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "300001", "FacilityName": "Reference Only"}],
            header=["FacilityNumber", "FacilityName"],
        )
        db_path = tmp_path / "does_not_exist.sqlite"

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
            only_facility_reference_rows=True,
        )

        rows = _read_csv(result.facility_overview_path)
        assert len(rows) == 1
        assert rows[0]["FacilityNumber"] == "300001"
        assert rows[0]["LoadedComplaintCount"] == "0"

    def test_substantiated_rows_limited_to_reference_facilities(
        self, tmp_path: Path
    ) -> None:
        """substantiated-complaints.csv excludes rows from non-reference facilities."""
        conn = _make_db()
        # Reference facility with substantiated complaint
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        # Unrelated facility with substantiated complaint — should be excluded
        _insert_facility(conn, fid="f2", fnum="999999")
        _insert_source_doc(conn, doc_id="d2", fid="f2", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c2", fid="f2", doc_id="d2",
            control_number="CR-002", finding="Substantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "100001", "FacilityName": "In Reference"}],
            header=["FacilityNumber", "FacilityName"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
            only_facility_reference_rows=True,
        )

        sub_rows = _read_csv(result.substantiated_complaints_path)
        facility_numbers = {r["FacilityNumber"] for r in sub_rows}
        assert "100001" in facility_numbers
        assert "999999" not in facility_numbers
        assert result.substantiated_complaint_row_count == 1

    def test_without_reference_csv_raises_clearly(self, tmp_path: Path) -> None:
        """Using only_facility_reference_rows without a CSV fails with a clear error."""
        import pytest

        with pytest.raises(FacilityReferenceFilterError) as exc_info:
            export_stakeholder_facility_overview(
                tmp_path / "does_not_exist.sqlite",
                tmp_path / "extracts",
                only_facility_reference_rows=True,
            )

        assert "reference" in str(exc_info.value).casefold()

    def test_default_behavior_unchanged_without_option(
        self, tmp_path: Path
    ) -> None:
        """Without only_facility_reference_rows, unrelated loaded facilities appear."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        _insert_facility(conn, fid="f2", fnum="999999")
        _insert_source_doc(conn, doc_id="d2", fid="f2", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c2", fid="f2", doc_id="d2",
            control_number="CR-002", finding="Unsubstantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "100001", "FacilityName": "In Reference"}],
            header=["FacilityNumber", "FacilityName"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
            # only_facility_reference_rows NOT set
        )

        rows = _read_csv(result.facility_overview_path)
        facility_numbers = {r["FacilityNumber"] for r in rows}
        # Both facilities present — default merges reference but does not filter
        assert "100001" in facility_numbers
        assert "999999" in facility_numbers
        assert result.facility_row_count == 2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_ref_csv(path: Path, rows: list[dict[str, str]], header: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _save_db(conn: sqlite3.Connection, path: Path) -> None:
    file_conn = sqlite3.connect(path)
    conn.backup(file_conn)
    file_conn.close()


def _read_csv(path: Path) -> list[dict[str, str]]:
    """Read CSV as list of dicts (excludes header row)."""
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def _read_csv_raw(path: Path) -> list[list[str]]:
    """Read CSV as list of lists (includes header row)."""
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.reader(f))


def _facility_overview_headers() -> tuple[str, ...]:
    from ccld_complaints.stakeholder_extract import _FACILITY_OVERVIEW_FIELDS
    return _FACILITY_OVERVIEW_FIELDS


def _substantiated_complaints_headers() -> tuple[str, ...]:
    from ccld_complaints.stakeholder_extract import _SUBSTANTIATED_COMPLAINTS_FIELDS
    return _SUBSTANTIATED_COMPLAINTS_FIELDS


def _insert_allegation(
    conn: sqlite3.Connection,
    *,
    aid: str,
    cid: str,
    allegation_text: str = "allegation text",
    category: str | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO allegations
            (allegation_id, complaint_id, allegation_text, allegation_category)
        VALUES (?, ?, ?, ?)
        """,
        (aid, cid, allegation_text, category),
    )


# ---------------------------------------------------------------------------
# Tests: complaint-records.csv
# ---------------------------------------------------------------------------


class TestComplaintRecordsExport:
    def _make_two_complaint_db(
        self, tmp_path: Path
    ) -> tuple[Path, Path]:
        """Return (db_path, extracts_root) with one substantiated + one unsubstantiated."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001", name="Alpha Care")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_source_doc(conn, doc_id="d2", fid="f1", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        _insert_complaint(
            conn, cid="c2", fid="f1", doc_id="d2",
            control_number="CR-002", finding="Unsubstantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)
        return db_path, tmp_path / "extracts"

    def test_complaint_records_includes_all_complaint_statuses(
        self, tmp_path: Path
    ) -> None:
        """complaint-records.csv has both substantiated and non-substantiated rows."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        rows = _read_csv(result.complaint_records_path)
        control_numbers = {r["ComplaintControlNumber"] for r in rows}
        assert "CR-001" in control_numbers
        assert "CR-002" in control_numbers
        assert result.complaint_record_row_count == 2

    def test_substantiated_complaints_csv_unchanged(self, tmp_path: Path) -> None:
        """substantiated-complaints.csv still contains only substantiated/equivalent."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        sub_rows = _read_csv(result.substantiated_complaints_path)
        assert len(sub_rows) == 1
        assert sub_rows[0]["ComplaintControlNumber"] == "CR-001"
        assert result.substantiated_complaint_row_count == 1

    def test_complaint_records_excludes_raw_narrative_text(
        self, tmp_path: Path
    ) -> None:
        """complaint-records.csv columns must not include allegation_text."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        raw = _read_csv_raw(result.complaint_records_path)
        headers = raw[0]
        prohibited = {"allegation_text", "AllegationText", "allegation text"}
        assert not prohibited.intersection(set(headers))
        # Also verify no row cell contains the raw string "allegation text"
        for data_row in raw[1:]:
            assert "allegation text" not in " ".join(data_row).casefold()

    def test_manifest_includes_complaint_record_row_count(
        self, tmp_path: Path
    ) -> None:
        """manifest.json has complaint_record_row_count and complaint-records.csv."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
        assert "complaint_record_row_count" in manifest
        assert manifest["complaint_record_row_count"] == 2
        assert "complaint-records.csv" in manifest["output_files"]

    def test_readme_explains_complaint_records(self, tmp_path: Path) -> None:
        """README.md mentions complaint-records.csv and its cautious limitations."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        readme = result.readme_path.read_text(encoding="utf-8")
        assert "complaint-records.csv" in readme
        assert "KeywordReviewCues" in readme
        assert "review-cue" in readme.casefold()
        assert "not a severity score" in readme.casefold()

    def test_missing_category_and_type_become_not_available(
        self, tmp_path: Path
    ) -> None:
        """Missing AllegationCategory and ComplaintType become 'not available'."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(
            conn, doc_id="d1", fid="f1", url="http://example.test/1"
        )
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")
        rows = _read_csv(result.complaint_records_path)
        assert rows[0]["AllegationCategory"] == UNKNOWN
        assert rows[0]["ComplaintType"] == UNKNOWN

    def test_finding_group_derived_from_source_finding(
        self, tmp_path: Path
    ) -> None:
        """FindingGroup values come from source-derived finding text only."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        for i, (doc_id, ctrl, finding) in enumerate([
            ("d1", "CR-001", "Substantiated"),
            ("d2", "CR-002", "Unsubstantiated"),
            ("d3", "CR-003", ""),
        ]):
            _insert_source_doc(
                conn, doc_id=doc_id, fid="f1",
                url=f"http://example.test/{i}",
            )
            _insert_complaint(
                conn, cid=f"c{i}", fid="f1", doc_id=doc_id,
                control_number=ctrl, finding=finding,
            )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")
        rows = _read_csv(result.complaint_records_path)
        by_ctrl = {r["ComplaintControlNumber"]: r["FindingGroup"] for r in rows}
        assert by_ctrl["CR-001"] == "SubstantiatedOrEquivalent"
        assert by_ctrl["CR-002"] == "NotSubstantiatedOrEquivalent"
        assert by_ctrl["CR-003"] == "UnknownOrMissing"

    def test_keyword_review_cue_from_finding(self, tmp_path: Path) -> None:
        """KeywordReviewCues fires when finding contains a review-cue keyword."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated - Abuse",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")
        rows = _read_csv(result.complaint_records_path)
        assert rows[0]["KeywordReviewCues"] == "Possible serious allegation topic"

    def test_keyword_review_cue_from_allegation_category(
        self, tmp_path: Path
    ) -> None:
        """KeywordReviewCues fires when allegation_category matches a keyword."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        _insert_allegation(
            conn, aid="a1", cid="c1", category="Physical Abuse",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")
        rows = _read_csv(result.complaint_records_path)
        assert rows[0]["KeywordReviewCues"] == "Possible serious allegation topic"

    def test_keyword_review_cue_not_available_when_no_match(
        self, tmp_path: Path
    ) -> None:
        """KeywordReviewCues is UNKNOWN when no keyword matches."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        _insert_allegation(conn, aid="a1", cid="c1", category="Licensing Violation")
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")
        rows = _read_csv(result.complaint_records_path)
        assert rows[0]["KeywordReviewCues"] == UNKNOWN

    def test_allegation_category_populated_from_allegations_table(
        self, tmp_path: Path
    ) -> None:
        """AllegationCategory in complaint-records.csv comes from allegations table."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        _insert_allegation(conn, aid="a1", cid="c1", category="Neglect")
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")
        rows = _read_csv(result.complaint_records_path)
        assert "Neglect" in rows[0]["AllegationCategory"]

    def test_complaint_records_in_xlsx(self, tmp_path: Path) -> None:
        """Complaint Records worksheet is included in the Excel workbook."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        assert "Complaint Records" in wb.sheetnames

    def test_complaint_records_respects_only_facility_reference_rows(
        self, tmp_path: Path
    ) -> None:
        """complaint-records.csv excludes non-reference facilities when filter enabled."""
        conn = _make_db()
        # Reference facility
        _insert_facility(conn, fid="f1", fnum="100001", name="In Reference")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Unsubstantiated",
        )
        # Unrelated facility (not in reference)
        _insert_facility(conn, fid="f2", fnum="999999", name="Not In Reference")
        _insert_source_doc(conn, doc_id="d2", fid="f2", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c2", fid="f2", doc_id="d2",
            control_number="CR-002", finding="Substantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)

        ref_csv = tmp_path / "facilities.csv"
        _write_ref_csv(
            ref_csv,
            [{"FacilityNumber": "100001", "FacilityName": "In Reference"}],
            header=["FacilityNumber", "FacilityName"],
        )

        result = export_stakeholder_facility_overview(
            db_path, tmp_path / "extracts",
            facility_reference_csv=ref_csv,
            only_facility_reference_rows=True,
        )

        rows = _read_csv(result.complaint_records_path)
        facility_numbers = {r["FacilityNumber"] for r in rows}
        assert "100001" in facility_numbers
        assert "999999" not in facility_numbers
        assert result.complaint_record_row_count == 1


# ---------------------------------------------------------------------------
# Tests: _derive_finding_group and _derive_keyword_review_cues
# ---------------------------------------------------------------------------


class TestFindingGroupDerivation:
    def _group(self, finding: str | None) -> str:
        from ccld_complaints.stakeholder_extract import _derive_finding_group
        return _derive_finding_group(finding)

    def test_substantiated_is_substantiated_group(self) -> None:
        assert self._group("Substantiated") == "SubstantiatedOrEquivalent"

    def test_founded_is_substantiated_group(self) -> None:
        assert self._group("Founded") == "SubstantiatedOrEquivalent"

    def test_unsubstantiated_is_not_substantiated_group(self) -> None:
        assert self._group("Unsubstantiated") == "NotSubstantiatedOrEquivalent"

    def test_inconclusive_is_not_substantiated_group(self) -> None:
        assert self._group("Inconclusive") == "NotSubstantiatedOrEquivalent"

    def test_dismissed_is_not_substantiated_group(self) -> None:
        assert self._group("Dismissed") == "NotSubstantiatedOrEquivalent"

    def test_empty_string_is_unknown(self) -> None:
        assert self._group("") == "UnknownOrMissing"

    def test_none_is_unknown(self) -> None:
        assert self._group(None) == "UnknownOrMissing"

    def test_not_available_sentinel_is_unknown(self) -> None:
        assert self._group(UNKNOWN) == "UnknownOrMissing"


class TestKeywordReviewCueDerivation:
    def _cue(self, finding: str | None, categories: str | None) -> str:
        from ccld_complaints.stakeholder_extract import _derive_keyword_review_cues
        return _derive_keyword_review_cues(finding, categories)

    def test_abuse_in_finding_triggers_cue(self) -> None:
        assert self._cue("Substantiated - Abuse", None) == "Possible serious allegation topic"

    def test_neglect_in_category_triggers_cue(self) -> None:
        assert self._cue(None, "Neglect") == "Possible serious allegation topic"

    def test_no_keyword_match_returns_not_available(self) -> None:
        assert self._cue("Unsubstantiated", "Licensing Violation") == UNKNOWN

    def test_none_finding_and_category_returns_not_available(self) -> None:
        assert self._cue(None, None) == UNKNOWN

    def test_cue_is_not_a_severity_score(self) -> None:
        """Confirm the cue label does not use severity/risk language."""
        result = self._cue("Neglect", None)
        assert "severity" not in result.casefold()
        assert "risk" not in result.casefold()
        assert "verified" not in result.casefold()


# ---------------------------------------------------------------------------
# Tests: Excel workbook output
# ---------------------------------------------------------------------------


class TestExcelWorkbook:
    def _make_two_complaint_db(self, tmp_path: Path) -> tuple[Path, Path]:
        """Return (db_path, extracts_root) with one substantiated + one unsubstantiated."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001", name="Alpha Care")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_source_doc(conn, doc_id="d2", fid="f1", url="http://example.test/2")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        _insert_complaint(
            conn, cid="c2", fid="f1", doc_id="d2",
            control_number="CR-002", finding="Unsubstantiated",
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)
        return db_path, tmp_path / "extracts"

    def test_xlsx_file_is_generated(self, tmp_path: Path) -> None:
        """The export produces an XLSX file."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        assert result.xlsx_path.exists()
        assert result.xlsx_path.suffix == ".xlsx"

    def test_zip_file_is_not_generated(self, tmp_path: Path) -> None:
        """No ZIP file is created as the final stakeholder package."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        zip_files = list(result.output_dir.glob("*.zip"))
        assert zip_files == [], f"Unexpected ZIP file(s): {zip_files}"

    def test_worksheet_order(self, tmp_path: Path) -> None:
        """Workbook sheet order: README, Summary, Facility Review Cues, Facility Overview,
        Substantiated Complaints, Complaint Records, Manifest."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        assert wb.sheetnames == [
            "README",
            "Summary",
            "Facility Review Cues",
            "Facility Overview",
            "Substantiated Complaints",
            "Complaint Records",
            "Manifest",
        ]

    def test_readme_worksheet_has_no_line_number_column(self, tmp_path: Path) -> None:
        """README worksheet column A must not contain bare numeric row labels."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["README"]
        col_a_values = [
            str(ws.cell(row=i, column=1).value or "")
            for i in range(1, ws.max_row + 1)
        ]
        for val in col_a_values:
            assert not val.strip().isdigit(), (
                f"README column A contains a bare numeric row label: {val!r}"
            )

    def test_readme_worksheet_includes_first_time_user_context(
        self, tmp_path: Path
    ) -> None:
        """README worksheet contains helpful first-time-user guidance."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["README"]
        all_text = " ".join(
            str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        ).casefold()
        assert "source of record" in all_text
        assert "legal conclusions" in all_text
        assert "how to use" in all_text
        assert "limitations" in all_text
        assert "loaded" in all_text

    def test_manifest_worksheet_is_key_value(self, tmp_path: Path) -> None:
        """Manifest worksheet uses key/value rows with display labels, not a raw JSON blob."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Manifest"]
        col_a_values = [
            str(ws.cell(row=i, column=1).value or "")
            for i in range(1, ws.max_row + 1)
        ]
        # Display label 'Generated' replaces snake_case 'generated_at'.
        assert any("Generated" in v for v in col_a_values)
        assert not any("generated_at" in v for v in col_a_values), (
            "Manifest should display 'Generated' not snake_case 'generated_at'"
        )
        assert not any(v.strip().startswith("{") for v in col_a_values), (
            "Manifest sheet should not contain raw JSON blobs"
        )

    def test_manifest_worksheet_includes_complaint_record_row_count(
        self, tmp_path: Path
    ) -> None:
        """Manifest worksheet includes Complaint Record Row Count display label."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Manifest"]
        col_a_values = [
            str(ws.cell(row=i, column=1).value or "")
            for i in range(1, ws.max_row + 1)
        ]
        assert "Complaint Record Row Count" in col_a_values

    def test_data_worksheet_row_counts_match(self, tmp_path: Path) -> None:
        """Data worksheet row counts (excluding header) match result row counts."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws_fo = wb["Facility Overview"]
        assert ws_fo.max_row - 1 == result.facility_row_count

        ws_cr = wb["Complaint Records"]
        assert ws_cr.max_row - 1 == result.complaint_record_row_count

    def test_complaint_records_more_than_substantiated(self, tmp_path: Path) -> None:
        """Complaint Records worksheet has more data rows than Substantiated Complaints."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        cr_rows = wb["Complaint Records"].max_row - 1
        sub_rows = wb["Substantiated Complaints"].max_row - 1
        assert cr_rows > sub_rows

    def test_complaint_records_worksheet_includes_all_finding_groups(
        self, tmp_path: Path
    ) -> None:
        """Complaint Records worksheet includes both substantiated and non-substantiated."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Complaint Records"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        ctrl_col = header.index("Complaint Control Number") + 1
        control_numbers = {
            ws.cell(row=r, column=ctrl_col).value
            for r in range(2, ws.max_row + 1)
        }
        assert "CR-001" in control_numbers
        assert "CR-002" in control_numbers

    def test_no_raw_narrative_in_xlsx(self, tmp_path: Path) -> None:
        """No raw allegation narrative text appears in any XLSX worksheet."""
        conn = _make_db()
        _insert_facility(conn, fid="f1", fnum="100001")
        _insert_source_doc(conn, doc_id="d1", fid="f1", url="http://example.test/1")
        _insert_complaint(
            conn, cid="c1", fid="f1", doc_id="d1",
            control_number="CR-001", finding="Substantiated",
        )
        conn.execute(
            """
            INSERT INTO allegations (allegation_id, complaint_id, allegation_text)
            VALUES ('a1', 'c1', 'XLSX_SENTINEL_NARRATIVE_SHOULD_NOT_APPEAR')
            """
        )
        conn.commit()
        db_path = tmp_path / "test.sqlite"
        _save_db(conn, db_path)
        result = export_stakeholder_facility_overview(db_path, tmp_path / "extracts")

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    assert "XLSX_SENTINEL_NARRATIVE_SHOULD_NOT_APPEAR" not in str(
                        cell.value or ""
                    ), f"Narrative text found in sheet {sheet_name!r}"

    def test_data_worksheets_freeze_top_row(self, tmp_path: Path) -> None:
        """Wide data worksheets freeze the header row and first two identifying columns
        (freeze_panes == 'C2'), keeping Facility Number and Facility Name visible."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in [
            "Facility Overview",
            "Substantiated Complaints",
            "Complaint Records",
        ]:
            ws = wb[sheet_name]
            assert ws.freeze_panes == "C2", (
                f"Sheet {sheet_name!r} should freeze top row and two left columns "
                f"(freeze_panes='C2'), got {ws.freeze_panes!r}"
            )

    def test_source_url_cells_have_hyperlinks(self, tmp_path: Path) -> None:
        """Source URL column cells in data sheets carry hyperlink objects for http URLs."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in ["Substantiated Complaints", "Complaint Records"]:
            ws = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            assert "Source URL" in header, f"{sheet_name!r} missing 'Source URL' column"
            url_col = header.index("Source URL") + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=url_col)
                val = str(cell.value or "")
                if val.startswith("http"):
                    assert cell.hyperlink is not None, (
                        f"Source URL cell in {sheet_name!r} row {row_idx} has no hyperlink"
                    )

    def test_review_cues_column_has_wrap_text(self, tmp_path: Path) -> None:
        """Review Cues and Suggested Next Step columns in Facility Review Cues have wrap_text."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Facility Review Cues"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for col_name in ("Review Cues", "Suggested Next Step"):
            assert col_name in header, f"Facility Review Cues missing {col_name!r} column"
            col_idx = header.index(col_name) + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                assert cell.alignment.wrap_text, (
                    f"{col_name!r} cell in Facility Review Cues row {row_idx} missing wrap_text"
                )

    def test_manifest_worksheet_has_auto_filter(self, tmp_path: Path) -> None:
        """Manifest worksheet has an auto_filter set."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Manifest"]
        assert ws.auto_filter.ref is not None, "Manifest sheet should have auto_filter"

    def test_readme_has_tabs_included_section(self, tmp_path: Path) -> None:
        """README worksheet contains a 'TABS INCLUDED' section listing all tabs."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["README"]
        all_text = " ".join(
            str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        ).casefold()
        assert "tabs included" in all_text
        assert "facility review cues" in all_text
        assert "facility overview" in all_text
        assert "substantiated complaints" in all_text
        assert "complaint records" in all_text
        assert "manifest" in all_text

    def test_readme_has_counts_and_coverage_guidance(self, tmp_path: Path) -> None:
        """README worksheet contains counts/coverage caution language."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["README"]
        all_text = " ".join(
            str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        ).casefold()
        assert "counts and coverage" in all_text
        assert "zero" in all_text or "does not prove" in all_text or "does not mean" in all_text

    def test_summary_worksheet_exists_after_readme(self, tmp_path: Path) -> None:
        """Summary worksheet is the second sheet, immediately after README."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        assert wb.sheetnames[1] == "Summary"

    def test_summary_worksheet_has_source_derived_metric_labels(
        self, tmp_path: Path
    ) -> None:
        """Summary worksheet contains source-derived count labels."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Summary"]
        all_text = " ".join(
            str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        ).casefold()
        assert "source-derived" in all_text
        assert "loaded complaint records" in all_text
        assert "substantiated" in all_text
        assert "finding group" in all_text
        assert "keyword review cue" in all_text
        assert "facility status" in all_text

    def test_summary_worksheet_has_cautious_wording(self, tmp_path: Path) -> None:
        """Summary worksheet contains cautious source-of-record and non-conclusion wording."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Summary"]
        all_text = " ".join(
            str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        ).casefold()
        assert "source of record" in all_text
        assert "does not make legal conclusions" in all_text or "legal conclusions" in all_text
        assert "source-completeness" in all_text
        assert "does not prove" in all_text or "zero" in all_text

    def test_summary_worksheet_has_no_prohibited_wording(self, tmp_path: Path) -> None:
        """Summary worksheet does not contain positive-claim wording that would imply
        risk scores, rankings, or verified conclusions."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Summary"]
        all_text = " ".join(
            str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        ).casefold()
        # These substrings only appear as positive claims, not in cautious disclaimers.
        assert "is a risk score" not in all_text
        assert "is a severity score" not in all_text
        assert "ranked by" not in all_text

    def test_all_worksheets_have_tab_colors(self, tmp_path: Path) -> None:
        """All five generated worksheets have a tab color set."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tab_color = ws.sheet_properties.tabColor
            assert tab_color is not None, (
                f"Sheet {sheet_name!r} has no tab color set"
            )
            assert str(tab_color.rgb) not in ("00000000", ""), (
                f"Sheet {sheet_name!r} tab color is empty/transparent"
            )

    def test_data_worksheet_headers_have_fill(self, tmp_path: Path) -> None:
        """Header row cells in data sheets have a background fill."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in [
            "Facility Overview",
            "Substantiated Complaints",
            "Complaint Records",
        ]:
            ws = wb[sheet_name]
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                fill = cell.fill
                assert fill is not None and fill.fill_type not in (None, "none"), (
                    f"Header cell {sheet_name!r} col {col_idx} has no fill"
                )

    def test_readme_title_row_has_dark_fill(self, tmp_path: Path) -> None:
        """README row 1 cell A1 has a non-default background fill (title styling)."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["README"]
        title_cell = ws.cell(row=1, column=1)
        fill = title_cell.fill
        assert fill is not None and fill.fill_type not in (None, "none"), (
            "README title cell A1 should have a background fill"
        )

    def test_manifest_header_row_has_fill(self, tmp_path: Path) -> None:
        """Manifest worksheet header row cells have a background fill."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Manifest"]
        for col_idx in (1, 2):
            cell = ws.cell(row=1, column=col_idx)
            fill = cell.fill
            assert fill is not None and fill.fill_type not in (None, "none"), (
                f"Manifest header cell col {col_idx} has no fill"
            )

    def test_facility_review_cues_has_one_row_per_facility(
        self, tmp_path: Path
    ) -> None:
        """Facility Review Cues has exactly one data row per facility overview row."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws_frc = wb["Facility Review Cues"]
        ws_fo = wb["Facility Overview"]
        # Both have same number of data rows (header excluded).
        assert ws_frc.max_row - 1 == ws_fo.max_row - 1
        assert ws_frc.max_row - 1 == result.facility_row_count

    def test_facility_review_cues_columns_are_present(self, tmp_path: Path) -> None:
        """Facility Review Cues contains expected display column headers."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Facility Review Cues"]
        header = [
            ws.cell(row=1, column=c).value
            for c in range(1, ws.max_column + 1)
        ]
        for col in (
            "Facility Number",
            "Facility Name",
            "Status",
            "County",
            "Loaded Complaint Count",
            "Substantiated Count",
            "Review Cues",
            "Suggested Next Step",
        ):
            assert col in header, f"Facility Review Cues missing column {col!r}"
        # Limitations removed from XLSX display.
        assert "Limitations" not in header, (
            "Limitations column should not appear in Facility Review Cues XLSX"
        )

    def test_facility_review_cues_review_cue_labels_are_deterministic(
        self, tmp_path: Path
    ) -> None:
        """Review Cues values in Facility Review Cues use cautious source-derived labels."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Facility Review Cues"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        cue_col = header.index("Review Cues") + 1
        cue_texts = " ".join(
            str(ws.cell(row=r, column=cue_col).value or "")
            for r in range(2, ws.max_row + 1)
        ).casefold()
        # Facility f1 has 1 substantiated + 1 unsubstantiated — expect these labels.
        assert "substantiated complaint records loaded" in cue_texts
        assert "multiple loaded complaint records" in cue_texts
        # Must not contain risk/ranking/score language.
        assert "risk" not in cue_texts
        assert "priority" not in cue_texts
        assert "ranked" not in cue_texts

    def test_facility_review_cues_suggested_next_step_wording(
        self, tmp_path: Path
    ) -> None:
        """Suggested Next Step values are cautious and do not imply verified conclusions."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Facility Review Cues"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        step_col = header.index("Suggested Next Step") + 1
        step_texts = " ".join(
            str(ws.cell(row=r, column=step_col).value or "")
            for r in range(2, ws.max_row + 1)
        ).casefold()
        # Cautious wording must be present.
        assert "verify" in step_texts or "confirm" in step_texts
        assert "ccld portal" in step_texts or "public" in step_texts
        # Must not imply a definitive conclusion.
        assert "confirmed" not in step_texts
        assert "proves" not in step_texts
        assert "risk" not in step_texts

    def test_summary_suppresses_all_not_available_cue_counts(
        self, tmp_path: Path
    ) -> None:
        """Summary shows a suppression note when all keyword cues are not available."""
        # _make_two_complaint_db uses findings 'Substantiated'/'Unsubstantiated',
        # neither of which matches keyword_review_cue keywords — all cues are UNKNOWN.
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Summary"]
        # Collect all column-A cell values (these are the label/section cells).
        col_a_values = [
            str(ws.cell(row=r, column=1).value or "")
            for r in range(1, ws.max_row + 1)
        ]
        col_a_lower = [v.casefold() for v in col_a_values]
        # Suppression note must appear somewhere in column A.
        assert any("not available in this extract" in v for v in col_a_lower), (
            "Expected suppression note not found in Summary column A"
        )
        # The count-row label for a cue match must NOT appear as a standalone row label
        # (it would appear as a regular row if counts were emitted instead of suppressed).
        assert not any(
            "possible serious allegation topic" in v
            and not any(
                keyword in v
                for keyword in ("signals", "not a", "review aid", "note:")
            )
            for v in col_a_lower
        ), "Count row for 'Possible serious allegation topic' should be suppressed"

    def test_data_cells_with_wrap_have_top_left_alignment(self, tmp_path: Path) -> None:
        """Wrapped Review Cues and Suggested Next Step cells are top-left aligned."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Facility Review Cues"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for col_name in ("Review Cues", "Suggested Next Step"):
            col_idx = header.index(col_name) + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                assert cell.alignment.horizontal == "left", (
                    f"Facility Review Cues {col_name!r} row {row_idx} horizontal != 'left'"
                )
                assert cell.alignment.vertical == "top", (
                    f"Facility Review Cues {col_name!r} row {row_idx} vertical != 'top'"
                )
                assert cell.alignment.wrap_text, (
                    f"Facility Review Cues {col_name!r} row {row_idx} missing wrap_text"
                )

    def test_all_table_based_sheets_have_freeze_and_filter(
        self, tmp_path: Path
    ) -> None:
        """Every table-based generated worksheet has freeze panes and an auto_filter."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        table_sheets = [
            "Facility Review Cues",
            "Facility Overview",
            "Substantiated Complaints",
            "Complaint Records",
            "Manifest",
        ]
        for sheet_name in table_sheets:
            ws = wb[sheet_name]
            assert ws.freeze_panes is not None and ws.freeze_panes != "", (
                f"Sheet {sheet_name!r} has no freeze panes"
            )
            assert ws.auto_filter.ref is not None, (
                f"Sheet {sheet_name!r} has no auto_filter"
            )

    def test_summary_sheet_has_freeze_and_filter(self, tmp_path: Path) -> None:
        """Summary worksheet has freeze panes and an auto_filter."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Summary"]
        assert ws.freeze_panes is not None and ws.freeze_panes != "", (
            "Summary sheet has no freeze panes"
        )
        assert ws.auto_filter.ref is not None, "Summary sheet has no auto_filter"

    def test_non_wrap_data_cells_are_top_left_aligned(self, tmp_path: Path) -> None:
        """Non-wrap data cells (Facility Number) in data sheets are left/top aligned."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in ["Facility Overview", "Facility Review Cues", "Complaint Records"]:
            ws = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            assert "Facility Number" in header, f"{sheet_name!r} missing 'Facility Number'"
            fn_col = header.index("Facility Number") + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=fn_col)
                if cell.value is None:
                    continue
                assert cell.alignment.horizontal == "left", (
                    f"{sheet_name!r} Facility Number row {row_idx} horizontal != 'left'"
                )
                assert cell.alignment.vertical == "top", (
                    f"{sheet_name!r} Facility Number row {row_idx} vertical != 'top'"
                )

    def test_facility_review_cues_limitations_column_absent_from_xlsx(
        self, tmp_path: Path
    ) -> None:
        """Limitations column is intentionally absent from the Facility Review Cues XLSX tab.
        Cautious source-of-record wording is instead centralised in README and Summary."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Facility Review Cues"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Limitations" not in header, (
            "Limitations column should be absent from Facility Review Cues XLSX tab"
        )

    def test_xlsx_column_headers_use_display_names(self, tmp_path: Path) -> None:
        """XLSX data worksheets use stakeholder-friendly spaced column display names."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in ["Facility Overview", "Complaint Records", "Facility Review Cues"]:
            ws = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            assert "Facility Number" in header, f"{sheet_name!r}: missing 'Facility Number'"
            assert "Facility Name" in header, f"{sheet_name!r}: missing 'Facility Name'"
            assert "FacilityNumber" not in header, (
                f"{sheet_name!r}: internal 'FacilityNumber' should not appear as header"
            )
            assert "FacilityName" not in header, (
                f"{sheet_name!r}: internal 'FacilityName' should not appear as header"
            )
        ws_sub = wb["Substantiated Complaints"]
        sub_header = [ws_sub.cell(row=1, column=c).value for c in range(1, ws_sub.max_column + 1)]
        assert "Source URL" in sub_header, "Substantiated Complaints: missing 'Source URL'"
        assert "SourceUrl" not in sub_header, (
            "Substantiated Complaints: 'SourceUrl' should not appear as header"
        )

    def test_finding_group_uses_stakeholder_terms(self, tmp_path: Path) -> None:
        """Complaint Records XLSX uses 'Substantiated'/'Unsubstantiated' not internal codes."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Complaint Records"]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        fg_col = header.index("Finding Group") + 1
        fg_values = {
            ws.cell(row=r, column=fg_col).value
            for r in range(2, ws.max_row + 1)
        }
        assert "Substantiated" in fg_values, (
            "Complaint Records: expected 'Substantiated' in Finding Group column"
        )
        assert "Unsubstantiated" in fg_values, (
            "Complaint Records: expected 'Unsubstantiated' in Finding Group column"
        )
        for val in fg_values:
            assert val not in ("SubstantiatedOrEquivalent", "NotSubstantiatedOrEquivalent"), (
                f"Complaint Records: internal FindingGroup code {val!r} should not appear"
            )

    def test_summary_generated_timestamp_is_human_readable(self, tmp_path: Path) -> None:
        """Summary worksheet 'Generated' row shows a human-readable timestamp, not compact form."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Summary"]
        generated_val = ""
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Generated":
                generated_val = str(ws.cell(row=r, column=2).value or "")
                break
        assert generated_val, "Summary: 'Generated' row not found or value is empty"
        # Compact form contains 'T' between date and time; human-readable form must not.
        import re
        assert not re.match(r"\d{8}T\d{6}Z", generated_val), (
            f"Summary 'Generated' value looks like compact form: {generated_val!r}"
        )
        assert "UTC" in generated_val, (
            f"Summary 'Generated' value should contain 'UTC': {generated_val!r}"
        )

    def test_limitations_absent_from_all_xlsx_data_sheets(self, tmp_path: Path) -> None:
        """Limitations column is intentionally removed from all four XLSX data tabs."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in [
            "Facility Review Cues",
            "Facility Overview",
            "Substantiated Complaints",
            "Complaint Records",
        ]:
            ws = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            assert "Limitations" not in header, (
                f"{sheet_name!r}: Limitations column should be absent from XLSX display"
            )

    def test_readme_generated_timestamp_is_human_readable(self, tmp_path: Path) -> None:
        """README KEY DETAILS Generated value is human-readable, not compact form."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["README"]
        generated_val = ""
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Generated":
                generated_val = str(ws.cell(row=r, column=2).value or "")
                break
        assert generated_val, "README: 'Generated' key row not found"
        import re
        assert not re.match(r"\d{8}T\d{6}Z", generated_val), (
            f"README 'Generated' value looks like compact form: {generated_val!r}"
        )
        assert "UTC" in generated_val, (
            f"README 'Generated' value should contain 'UTC': {generated_val!r}"
        )

    def test_manifest_generated_timestamp_is_human_readable(self, tmp_path: Path) -> None:
        """Manifest Generated row shows human-readable timestamp, not compact form."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Manifest"]
        generated_val = ""
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Generated":
                generated_val = str(ws.cell(row=r, column=2).value or "")
                break
        assert generated_val, "Manifest: 'Generated' row not found"
        import re
        assert not re.match(r"\d{8}T\d{6}Z", generated_val), (
            f"Manifest 'Generated' value looks like compact form: {generated_val!r}"
        )
        assert "UTC" in generated_val, (
            f"Manifest 'Generated' value should contain 'UTC': {generated_val!r}"
        )

    def test_readme_and_summary_have_merged_narrative_rows(self, tmp_path: Path) -> None:
        """README and Summary have merged A:B cells for prose-only rows."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in ("README", "Summary"):
            ws = wb[sheet_name]
            merged = {r for r1, r2, c1, c2 in (
                (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
                for mr in ws.merged_cells.ranges
            ) if c1 == 1 and c2 == 2 and r1 == r2 for r in (r1,)}
            assert len(merged) > 0, (
                f"{sheet_name!r}: expected at least one narrative row merged across A:B"
            )

    def test_no_merged_cells_in_data_table_sheets(self, tmp_path: Path) -> None:
        """Data table worksheets (Facility Overview etc.) have no merged cells."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in [
            "Facility Overview",
            "Facility Review Cues",
            "Substantiated Complaints",
            "Complaint Records",
        ]:
            ws = wb[sheet_name]
            assert len(ws.merged_cells.ranges) == 0, (
                f"{sheet_name!r}: data table sheet should have no merged cells"
            )

    def test_no_internal_terminology_in_workbook(self, tmp_path: Path) -> None:
        """XLSX workbook cells do not contain internal or machine-style terminology."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        forbidden = (
            "SubstantiatedOrEquivalent",
            "NotSubstantiatedOrEquivalent",
            "substantiated/equivalent",
            "unsubstantiated/equivalent",
            "KeywordReviewCues",
            "SourceUrl",
            "SourceURL",
        )
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    val = str(cell.value or "")
                    for term in forbidden:
                        assert term not in val, (
                            f"Forbidden term {term!r} found in sheet {sheet_name!r} "
                            f"row {cell.row} col {cell.column}: {val!r}"
                        )

    def test_summary_finding_group_labels_are_stakeholder_friendly(
        self, tmp_path: Path
    ) -> None:
        """Summary finding group counts use 'Substantiated'/'Unsubstantiated' not codes."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Summary"]
        col_a_values = [
            str(ws.cell(row=r, column=1).value or "")
            for r in range(1, ws.max_row + 1)
        ]
        # The indented count rows for our two complaint findings should be friendly labels.
        assert any("Substantiated" in v for v in col_a_values), (
            "Summary: expected 'Substantiated' count row in finding group section"
        )
        assert any("Unsubstantiated" in v for v in col_a_values), (
            "Summary: expected 'Unsubstantiated' count row in finding group section"
        )
        # Internal codes must not appear.
        for v in col_a_values:
            assert "SubstantiatedOrEquivalent" not in v, (
                f"Summary: internal code 'SubstantiatedOrEquivalent' found in row: {v!r}"
            )
            assert "NotSubstantiatedOrEquivalent" not in v, (
                f"Summary: internal code 'NotSubstantiatedOrEquivalent' found in row: {v!r}"
            )

    def test_manifest_uses_display_labels(self, tmp_path: Path) -> None:
        """XLSX Manifest worksheet uses stakeholder display labels instead of snake_case keys."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        ws = wb["Manifest"]
        col_a_values = [
            str(ws.cell(row=i, column=1).value or "")
            for i in range(1, ws.max_row + 1)
        ]
        expected_labels = (
            "Generated",
            "Facility Row Count",
            "Substantiated Complaint Row Count",
            "Complaint Record Row Count",
        )
        for label in expected_labels:
            assert label in col_a_values, (
                f"Manifest: expected display label {label!r} in column A"
            )
        forbidden_keys = (
            "generated_at",
            "facility_row_count",
            "substantiated_complaint_row_count",
            "complaint_record_row_count",
        )
        for key in forbidden_keys:
            assert key not in col_a_values, (
                f"Manifest: snake_case key {key!r} should not appear in XLSX display"
            )

    def test_narrative_cells_have_no_leading_spaces(self, tmp_path: Path) -> None:
        """Merged narrative cells in README and Summary do not start with leading spaces."""
        db_path, extracts = self._make_two_complaint_db(tmp_path)
        result = export_stakeholder_facility_overview(db_path, extracts)

        wb = openpyxl.load_workbook(result.xlsx_path)
        for sheet_name in ("README", "Summary"):
            ws = wb[sheet_name]
            for mr in ws.merged_cells.ranges:
                if mr.min_col == 1 and mr.max_col == 2:
                    cell_val = str(ws.cell(row=mr.min_row, column=1).value or "")
                    assert not cell_val.startswith(" "), (
                        f"{sheet_name!r} merged narrative row {mr.min_row} "
                        f"has leading space: {cell_val[:30]!r}"
                    )

