"""Stakeholder facility overview extract generator.

Reads from the local SQLite database that is already populated by the
normal CCLD ingestion pipeline and writes a stakeholder-ready ZIP package
containing:

- facility-overview.csv        per-facility aggregation
- substantiated-complaints.csv individual substantiated/equivalent records
- README.md                    plain-language scope and limitations
- manifest.json                generation metadata and row counts

Definitions
-----------
Substantiated/equivalent uses the same conservative keyword matching as
the hosted cross-facility triage page: a finding/resolution/status value
is substantiated-equivalent when it contains "substantiated", "founded", or
"sustained" but does NOT contain "unsubstantiated" or "not substantiated".

Boundaries
----------
- Reads only. Never writes to the database.
- Does not include raw narrative text from allegations.
- Does not add risk scores, severity scores, or rankings.
- Does not make legal conclusions, facility-wide conclusions, or
  source-completeness claims.
- Produces empty CSVs with headers if the database is absent or empty
  rather than raising.
- Optional facility reference CSV enriches facility metadata but never
  implies that zero loaded complaints means no public complaints exist.
"""
from __future__ import annotations

import csv
import json
import sqlite3
import subprocess
from collections.abc import Mapping
from dataclasses import dataclass, field, replace
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ccld_complaints.facility_identity_exports import (
    FACILITY_EXPORT_IDENTITY_FIELDS,
    load_sqlite_facility_identity_projections,
    projected_facility_export_conflicts,
    projected_facility_export_context,
    projected_facility_export_text,
)
from ccld_complaints.hosted_app.facility_identity_projection import (
    FacilityIdentityProjection,
    FacilityProjectionField,
)

DEFAULT_STAKEHOLDER_EXTRACT_ROOT = Path("data/processed/stakeholder-extracts")
UNKNOWN = "not available"

# ---------------------------------------------------------------------------
# Facility reference CSV column aliases
# ---------------------------------------------------------------------------
# Each tuple lists accepted header spellings (case-insensitive, stripped) for
# that logical field. First match wins.
_FACILITY_NUMBER_ALIASES: tuple[str, ...] = (
    "facilitynumber",
    "facility number",
    "licensenumber",
    "license number",
    "license_number",
    "facility_number",
    "fac_nbr",
)
_FACILITY_NAME_ALIASES: tuple[str, ...] = (
    "facilityname",
    "facility name",
    "facility_name",
    "name",
)
_FACILITY_TYPE_ALIASES: tuple[str, ...] = (
    "facilitytype",
    "fac_type_desc",
    "facility type",
    "facility_type",
    "programtype",
    "program type",
    "program_type",
)
_STATUS_ALIASES: tuple[str, ...] = (
    "status",
    "facilitystatus",
    "facility status",
    "facility_status",
)
_CITY_ALIASES: tuple[str, ...] = ("city", "res_city")
_COUNTY_ALIASES: tuple[str, ...] = ("county",)

# ---------------------------------------------------------------------------
# Substantiated/equivalent matching — mirrors reviewer_ui._is_substantiated_equivalent
# ---------------------------------------------------------------------------
_SUBSTANTIATED_KEYWORDS: tuple[str, ...] = ("substantiated", "founded", "sustained")
_EXCLUDE_PREFIXES: tuple[str, ...] = ("unsubstantiated", "not substantiated")

# Keywords used for deterministic keyword review cues derived from non-narrative
# source-extracted fields only (finding, allegation_category).  Matches signal
# a review topic; they are NOT severity scores, risk scores, or verified findings.
_KEYWORD_REVIEW_CUE_KEYWORDS: tuple[str, ...] = (
    "sexual assault",
    "abuse",
    "neglect",
    "injury",
    "restraint",
    "runaway",
    "awol",
    "medication",
    "staff misconduct",
)


def is_substantiated_equivalent(value: str | None) -> bool:
    """Return True when *value* indicates a substantiated or equivalent finding."""
    if not value:
        return False
    normalized = " ".join(value.strip().casefold().split())
    if not normalized:
        return False
    for prefix in _EXCLUDE_PREFIXES:
        if prefix in normalized:
            return False
    return any(kw in normalized for kw in _SUBSTANTIATED_KEYWORDS)


# ---------------------------------------------------------------------------
# SQL queries (read-only, no narrative text)
# ---------------------------------------------------------------------------

_FACILITY_COMPLAINTS_SQL = """
SELECT
    f.external_facility_number  AS facility_number,
    f.facility_name,
    f.facility_type,
    f.status,
    f.county,
    c.complaint_id,
    c.complaint_control_number,
    c.complaint_received_date,
    c.report_date,
    c.finding,
    sd.source_url,
    sd.raw_sha256,
    sd.raw_path,
    CASE WHEN sd.raw_sha256 IS NOT NULL AND sd.raw_sha256 != '' THEN 1 ELSE 0 END
        AS has_traceability
FROM complaints c
JOIN facilities f ON f.facility_id = c.facility_id
JOIN source_documents sd ON sd.document_id = c.document_id
ORDER BY f.external_facility_number, c.complaint_received_date DESC
"""

_ALL_COMPLAINTS_SQL = """
SELECT
    f.external_facility_number  AS facility_number,
    c.complaint_control_number,
    c.complaint_received_date,
    c.report_date,
    c.finding,
    sd.document_type            AS complaint_type,
    sd.source_url,
    sd.raw_sha256,
    (
        SELECT GROUP_CONCAT(DISTINCT a.allegation_category)
        FROM allegations a
        WHERE a.complaint_id = c.complaint_id
          AND a.allegation_category IS NOT NULL
          AND a.allegation_category != ''
    ) AS allegation_categories
FROM complaints c
JOIN facilities f ON f.facility_id = c.facility_id
JOIN source_documents sd ON sd.document_id = c.document_id
ORDER BY f.external_facility_number, c.complaint_received_date DESC
"""

# Try to get city from facility if the column exists (populated in some setups)
_FACILITY_CITY_SQL = """
SELECT f.external_facility_number AS facility_number, NULL AS city
FROM facilities f
GROUP BY f.external_facility_number
"""

_FACILITY_SQL = """
SELECT
    f.external_facility_number AS facility_number,
    MAX(f.facility_name) AS facility_name,
    MAX(f.facility_type) AS facility_type,
    MAX(f.status) AS status,
    MAX(f.county) AS county
FROM facilities f
GROUP BY f.external_facility_number
"""


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class _FacilityRow:
    facility_number: str
    facility_name: str
    facility_type: str
    status: str
    city: str
    county: str
    loaded_complaint_count: int = 0
    substantiated_count: int = 0
    complaint_dates: list[str] = field(default_factory=list)
    source_traceability_ready: int = 0
    missing_traceability: int = 0
    facility_identity_context: str = "No selected source context"
    facility_identity_conflicts: str = "No conflicting source values"


@dataclass(frozen=True)
class _SubstantiatedRow:
    facility_number: str
    facility_name: str
    complaint_control_number: str
    complaint_received_date: str
    report_date: str
    finding_or_resolution: str
    source_url: str
    raw_hash_or_artifact_reference: str
    reviewer_detail_path: str
    limitations: str
    facility_identity_context: str = "No selected source context"
    facility_identity_conflicts: str = "No conflicting source values"


@dataclass(frozen=True)
class _ComplaintRecordRow:
    facility_number: str
    facility_name: str
    facility_type: str
    status: str
    city: str
    county: str
    complaint_control_number: str
    complaint_received_date: str
    report_date: str
    finding_or_resolution: str
    finding_group: str
    complaint_type: str
    allegation_category: str
    keyword_review_cues: str
    source_url: str
    raw_hash_or_artifact_reference: str
    reviewer_detail_path: str
    limitations: str
    facility_identity_context: str = "No selected source context"
    facility_identity_conflicts: str = "No conflicting source values"


@dataclass(frozen=True)
class StakeholderExtractResult:
    output_dir: Path
    facility_overview_path: Path
    substantiated_complaints_path: Path
    readme_path: Path
    manifest_path: Path
    xlsx_path: Path
    facility_row_count: int
    substantiated_complaint_row_count: int
    git_commit: str
    input_source: str
    limitations: str
    facility_reference_csv: str
    facility_reference_row_count: int
    facility_reference_matched_count: int
    only_facility_reference_rows: bool
    complaint_records_path: Path
    complaint_record_row_count: int


# ---------------------------------------------------------------------------
# Finding group and keyword review cue derivation
# ---------------------------------------------------------------------------

def _derive_finding_group(finding: str | None) -> str:
    """Classify a source-derived finding into a broad group label.

    Groups are derived only from source-derived finding/resolution text and
    do not constitute an independent legal determination.

    - SubstantiatedOrEquivalent:    finding contains substantiated/founded/sustained
    - NotSubstantiatedOrEquivalent: finding is present but not substantiated-equivalent
    - UnknownOrMissing:             finding is absent, empty, or UNKNOWN
    """
    if not finding or finding == UNKNOWN:
        return "UnknownOrMissing"
    if is_substantiated_equivalent(finding):
        return "SubstantiatedOrEquivalent"
    return "NotSubstantiatedOrEquivalent"


def _derive_keyword_review_cues(
    finding: str | None,
    allegation_categories: str | None,
) -> str:
    """Return a review-cue label when an extracted field matches a keyword.

    Checks only source-derived non-narrative fields: finding and
    allegation_category.  Does not inspect raw allegation text.

    This is a review cue, not a severity score, not a risk score, not a
    verified finding, and not a legal classification.

    Returns UNKNOWN when no keyword matches.
    """
    parts: list[str] = []
    if finding:
        parts.append(finding)
    if allegation_categories:
        parts.append(allegation_categories)
    haystack = " ".join(parts).casefold()
    if any(keyword in haystack for keyword in _KEYWORD_REVIEW_CUE_KEYWORDS):
        return "Possible serious allegation topic"
    return UNKNOWN


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_stakeholder_facility_overview(
    db_path: Path,
    output_root: Path = DEFAULT_STAKEHOLDER_EXTRACT_ROOT,
    facility_reference_csv: Path | None = None,
    only_facility_reference_rows: bool = False,
) -> StakeholderExtractResult:
    """Generate a stakeholder facility overview extract package.

    If *db_path* does not exist or contains no facilities, produces empty
    CSVs with correct headers, README, manifest, and ZIP without failing.

    If *facility_reference_csv* is provided its rows join the complaint-time
    observations in the shared governed facility projection. Facilities present
    in the reference CSV but absent from the database are included with zero
    complaint counts.
    Complaint-derived counts are always based only on loaded records.

    If *only_facility_reference_rows* is True, only facilities whose
    facility number appears in the reference CSV are included in output.
    Raises FacilityReferenceFilterError when used without a reference CSV.
    """
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    output_dir = output_root / timestamp
    output_dir.mkdir(parents=True, exist_ok=True)

    if only_facility_reference_rows and facility_reference_csv is None:
        raise FacilityReferenceFilterError(
            "--only-facility-reference-rows requires --facility-reference-csv to be set."
        )

    db_exists = db_path.exists()
    input_source = db_path.as_posix() if db_exists else f"{db_path.as_posix()} (not found)"
    limitations = _limitations_sentence()

    if db_exists:
        facility_rows, substantiated_rows, complaint_sql_rows = _read_db(db_path)
    else:
        facility_rows, substantiated_rows, complaint_sql_rows = [], [], []

    # Load every optional reference observation; duplicate Facility IDs are
    # reconciled by the shared projector rather than discarded by row order.
    ref_row_count = 0
    ref_matched_count = 0
    ref_csv_value = "none"
    ref_numbers: set[str] = set()
    ref_records: list[dict[str, str]] = []
    if facility_reference_csv is not None:
        ref_records = read_facility_reference_csv(facility_reference_csv)
        ref_row_count = len(ref_records)
        ref_csv_value = facility_reference_csv.as_posix()
        ref_numbers = {r["facility_number"] for r in ref_records}
    existing_facility_numbers = {row.facility_number for row in facility_rows}
    ref_matched_count = len(existing_facility_numbers & ref_numbers)
    if db_exists:
        with sqlite3.connect(db_path) as projection_connection:
            projection_connection.row_factory = sqlite3.Row
            projections = load_sqlite_facility_identity_projections(
                projection_connection,
                reference_records=ref_records,
            )
    else:
        with sqlite3.connect(":memory:") as projection_connection:
            projection_connection.row_factory = sqlite3.Row
            projections = load_sqlite_facility_identity_projections(
                projection_connection,
                reference_records=ref_records,
            )
    facility_rows = _facility_rows_from_projections(facility_rows, projections)
    substantiated_rows = _substantiated_rows_from_projections(
        substantiated_rows,
        projections,
    )

    # Build enriched facility map (used to enrich complaint record rows)
    all_facility_map: dict[str, _FacilityRow] = {
        row.facility_number: row for row in facility_rows
    }

    # Build complaint record rows from all loaded complaints, enriched by facility map
    complaint_record_rows = _build_complaint_records(complaint_sql_rows, all_facility_map)

    # Apply reference-only filter when requested
    if only_facility_reference_rows:
        facility_rows = [
            r for r in facility_rows if r.facility_number in ref_numbers
        ]
        substantiated_rows = [
            r for r in substantiated_rows if r.facility_number in ref_numbers
        ]
        complaint_record_rows = [
            r for r in complaint_record_rows if r.facility_number in ref_numbers
        ]

    facility_overview_path = output_dir / "facility-overview.csv"
    substantiated_complaints_path = output_dir / "substantiated-complaints.csv"
    complaint_records_path = output_dir / "complaint-records.csv"
    readme_path = output_dir / "README.md"
    manifest_path = output_dir / "manifest.json"
    xlsx_name = f"stakeholder-facility-overview-{timestamp}.xlsx"
    xlsx_path = output_dir / xlsx_name

    _write_facility_overview_csv(facility_overview_path, facility_rows)
    _write_substantiated_complaints_csv(substantiated_complaints_path, substantiated_rows)
    _write_complaint_records_csv(complaint_records_path, complaint_record_rows)
    readme_path.write_text(_readme_text(), encoding="utf-8")

    git_commit = _get_git_commit()
    manifest = _build_manifest(
        generated_at=timestamp,
        input_source=input_source,
        facility_row_count=len(facility_rows),
        substantiated_complaint_row_count=len(substantiated_rows),
        complaint_record_row_count=len(complaint_record_rows),
        git_commit=git_commit,
        output_files=[
            "facility-overview.csv",
            "substantiated-complaints.csv",
            "complaint-records.csv",
            "README.md",
            "manifest.json",
            xlsx_name,
        ],
        limitations=limitations,
        facility_reference_csv=ref_csv_value,
        facility_reference_row_count=ref_row_count,
        facility_reference_matched_count=ref_matched_count,
        only_facility_reference_rows=only_facility_reference_rows,
    )
    manifest_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    _write_xlsx_workbook(
        xlsx_path,
        facility_rows=facility_rows,
        substantiated_rows=substantiated_rows,
        complaint_record_rows=complaint_record_rows,
        manifest=manifest,
    )

    return StakeholderExtractResult(
        output_dir=output_dir,
        facility_overview_path=facility_overview_path,
        substantiated_complaints_path=substantiated_complaints_path,
        complaint_records_path=complaint_records_path,
        readme_path=readme_path,
        manifest_path=manifest_path,
        xlsx_path=xlsx_path,
        facility_row_count=len(facility_rows),
        substantiated_complaint_row_count=len(substantiated_rows),
        complaint_record_row_count=len(complaint_record_rows),
        git_commit=git_commit,
        input_source=input_source,
        limitations=limitations,
        facility_reference_csv=ref_csv_value,
        facility_reference_row_count=ref_row_count,
        facility_reference_matched_count=ref_matched_count,
        only_facility_reference_rows=only_facility_reference_rows,
    )


# ---------------------------------------------------------------------------
# Database reads
# ---------------------------------------------------------------------------

def _read_db(
    db_path: Path,
) -> tuple[list[_FacilityRow], list[_SubstantiatedRow], list[sqlite3.Row]]:
    """Read facility and complaint data from SQLite and aggregate."""
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        try:
            raw_rows = list(conn.execute(_FACILITY_COMPLAINTS_SQL).fetchall())
        except sqlite3.OperationalError:
            # Table does not exist yet — treat as empty
            return [], [], []
        try:
            complaint_sql_rows: list[sqlite3.Row] = list(
                conn.execute(_ALL_COMPLAINTS_SQL).fetchall()
            )
        except sqlite3.OperationalError:
            complaint_sql_rows = []

    facility_rows, substantiated_rows = _aggregate(raw_rows)
    return facility_rows, substantiated_rows, complaint_sql_rows


def _aggregate(
    raw_rows: list[sqlite3.Row],
) -> tuple[list[_FacilityRow], list[_SubstantiatedRow]]:
    facility_map: dict[str, _FacilityRow] = {}
    substantiated_rows: list[_SubstantiatedRow] = []

    for row in raw_rows:
        fnum = _str(row, "facility_number") or UNKNOWN
        if fnum not in facility_map:
            facility_map[fnum] = _FacilityRow(
                facility_number=fnum,
                facility_name=_str(row, "facility_name") or UNKNOWN,
                facility_type=_str(row, "facility_type") or UNKNOWN,
                status=_str(row, "status") or UNKNOWN,
                city=UNKNOWN,
                county=_str(row, "county") or UNKNOWN,
            )
        frow = facility_map[fnum]
        frow.loaded_complaint_count += 1

        has_trace = int(row["has_traceability"] or 0)
        if has_trace:
            frow.source_traceability_ready += 1
        else:
            frow.missing_traceability += 1

        received = _str(row, "complaint_received_date")
        if received:
            frow.complaint_dates.append(received)

        finding = _str(row, "finding")
        if is_substantiated_equivalent(finding):
            frow.substantiated_count += 1
            control_number = _str(row, "complaint_control_number") or UNKNOWN
            source_url = _str(row, "source_url") or UNKNOWN
            raw_hash = _str(row, "raw_sha256") or UNKNOWN
            detail_path = (
                f"/reviewer/records/detail?source_record_key="
                f"complaint%3Accld%3Acomplaint%3A{control_number}"
                if control_number != UNKNOWN
                else UNKNOWN
            )
            substantiated_rows.append(
                _SubstantiatedRow(
                    facility_number=fnum,
                    facility_name=frow.facility_name,
                    complaint_control_number=control_number,
                    complaint_received_date=_str(row, "complaint_received_date") or UNKNOWN,
                    report_date=_str(row, "report_date") or UNKNOWN,
                    finding_or_resolution=finding or UNKNOWN,
                    source_url=source_url,
                    raw_hash_or_artifact_reference=raw_hash,
                    reviewer_detail_path=detail_path,
                    limitations=_limitations_sentence(),
                )
            )

    facility_rows: list[_FacilityRow] = sorted(
        facility_map.values(), key=lambda f: f.facility_number
    )
    substantiated_rows.sort(
        key=lambda r: (r.facility_number, r.complaint_received_date)
    )
    return facility_rows, substantiated_rows


def _str(row: sqlite3.Row, key: str) -> str:
    val = row[key] if key in row.keys() else None
    if val is None:
        return ""
    return str(val).strip()


def _build_complaint_records(
    complaint_sql_rows: list[sqlite3.Row],
    facility_map: dict[str, _FacilityRow],
) -> list[_ComplaintRecordRow]:
    """Build complaint record rows from raw SQL rows, enriched by facility_map.

    Only source-derived non-narrative fields are used.  Raw allegation text
    is never read or included.
    """
    rows: list[_ComplaintRecordRow] = []
    for row in complaint_sql_rows:
        fnum = _str(row, "facility_number") or UNKNOWN
        frow = facility_map.get(fnum)
        facility_name = frow.facility_name if frow else UNKNOWN
        facility_type = frow.facility_type if frow else UNKNOWN
        status = frow.status if frow else UNKNOWN
        city = frow.city if frow else UNKNOWN
        county = frow.county if frow else UNKNOWN

        finding = _str(row, "finding")
        allegation_categories = _str(row, "allegation_categories")
        control_number = _str(row, "complaint_control_number") or UNKNOWN

        detail_path = (
            "/reviewer/records/detail?source_record_key="
            f"complaint%3Accld%3Acomplaint%3A{control_number}"
            if control_number != UNKNOWN
            else UNKNOWN
        )

        rows.append(
            _ComplaintRecordRow(
                facility_number=fnum,
                facility_name=facility_name,
                facility_type=facility_type,
                status=status,
                city=city,
                county=county,
                complaint_control_number=control_number,
                complaint_received_date=_str(row, "complaint_received_date") or UNKNOWN,
                report_date=_str(row, "report_date") or UNKNOWN,
                finding_or_resolution=finding or UNKNOWN,
                finding_group=_derive_finding_group(finding if finding else None),
                complaint_type=_str(row, "complaint_type") or UNKNOWN,
                allegation_category=allegation_categories or UNKNOWN,
                keyword_review_cues=_derive_keyword_review_cues(
                    finding if finding else None,
                    allegation_categories if allegation_categories else None,
                ),
                source_url=_str(row, "source_url") or UNKNOWN,
                raw_hash_or_artifact_reference=_str(row, "raw_sha256") or UNKNOWN,
                reviewer_detail_path=detail_path,
                limitations=_limitations_sentence(),
                facility_identity_context=(
                    frow.facility_identity_context if frow else "No selected source context"
                ),
                facility_identity_conflicts=(
                    frow.facility_identity_conflicts
                    if frow
                    else "No conflicting source values"
                ),
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Facility reference CSV reader
# ---------------------------------------------------------------------------

class FacilityReferenceError(ValueError):
    """Raised when the facility reference CSV lacks a usable facility number column."""


class FacilityReferenceFilterError(ValueError):
    """Raised when only_facility_reference_rows is requested without a reference CSV."""


def _match_alias(
    headers: list[str], aliases: tuple[str, ...]
) -> str | None:
    """Return the first header that matches any alias (case-insensitive stripped)."""
    lower_headers = [h.strip().casefold() for h in headers]
    for alias in aliases:
        for i, lh in enumerate(lower_headers):
            if lh == alias:
                return headers[i]
    return None


def read_facility_reference_csv(
    csv_path: Path,
) -> list[dict[str, str]]:
    """Read a facility reference CSV and return normalised row dicts.

    Each returned dict has canonical keys: facility_number, facility_name,
    facility_type, status, city, county.  Unknown / missing values are the
    empty string.

    Raises FacilityReferenceError if no usable facility number column is
    found. Duplicate Facility IDs remain separate observations for the shared
    governed projection to reconcile without an input-order winner.
    """
    with csv_path.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        raw_rows = list(reader)
        headers = list(reader.fieldnames or [])

    if not headers and raw_rows:
        headers = list(raw_rows[0].keys())

    fnum_col = _match_alias(headers, _FACILITY_NUMBER_ALIASES)
    if fnum_col is None:
        raise FacilityReferenceError(
            f"Facility reference CSV '{csv_path}' does not contain a recognised "
            "facility/license number column. "
            f"Accepted column names (case-insensitive): "
            f"{', '.join(_FACILITY_NUMBER_ALIASES)}"
        )

    name_col = _match_alias(headers, _FACILITY_NAME_ALIASES)
    type_col = _match_alias(headers, _FACILITY_TYPE_ALIASES)
    status_col = _match_alias(headers, _STATUS_ALIASES)
    city_col = _match_alias(headers, _CITY_ALIASES)
    county_col = _match_alias(headers, _COUNTY_ALIASES)

    result: list[dict[str, str]] = []
    for raw in raw_rows:
        fnum = (raw.get(fnum_col) or "").strip()
        if not fnum:
            continue  # skip rows with no facility number
        result.append({
            "facility_number": fnum,
            "facility_name": (raw.get(name_col) or "").strip() if name_col else "",
            "facility_type": (raw.get(type_col) or "").strip() if type_col else "",
            "status": (raw.get(status_col) or "").strip() if status_col else "",
            "city": (raw.get(city_col) or "").strip() if city_col else "",
            "county": (raw.get(county_col) or "").strip() if county_col else "",
        })
    return result


def _facility_rows_from_projections(
    complaint_rows: list[_FacilityRow],
    projections: Mapping[str, FacilityIdentityProjection],
) -> list[_FacilityRow]:
    existing = {row.facility_number: row for row in complaint_rows}
    rows: list[_FacilityRow] = []
    for facility_number, projection in sorted(projections.items()):
        prior = existing.get(facility_number, _FacilityRow(
            facility_number=facility_number,
            facility_name=UNKNOWN,
            facility_type=UNKNOWN,
            status=UNKNOWN,
            city=UNKNOWN,
            county=UNKNOWN,
        ))
        rows.append(
            replace(
                prior,
                facility_number=projection.public_facility_id,
                facility_name=projected_facility_export_text(
                    projection, FacilityProjectionField.FACILITY_NAME
                ),
                facility_type=projected_facility_export_text(
                    projection, FacilityProjectionField.FACILITY_TYPE
                ),
                status=projected_facility_export_text(
                    projection, FacilityProjectionField.STATUS
                ),
                city=projected_facility_export_text(
                    projection, FacilityProjectionField.CITY
                ),
                county=projected_facility_export_text(
                    projection, FacilityProjectionField.COUNTY
                ),
                facility_identity_context=projected_facility_export_context(
                    projection,
                    FACILITY_EXPORT_IDENTITY_FIELDS,
                ),
                facility_identity_conflicts=projected_facility_export_conflicts(
                    projection,
                    FACILITY_EXPORT_IDENTITY_FIELDS,
                ),
            )
        )
    return rows


def _substantiated_rows_from_projections(
    rows: list[_SubstantiatedRow],
    projections: Mapping[str, FacilityIdentityProjection],
) -> list[_SubstantiatedRow]:
    projected: list[_SubstantiatedRow] = []
    for row in rows:
        projection = projections.get(row.facility_number)
        if projection is None:
            projected.append(row)
            continue
        projected.append(
            replace(
                row,
                facility_number=projection.public_facility_id,
                facility_name=projected_facility_export_text(
                    projection,
                    FacilityProjectionField.FACILITY_NAME,
                ),
                facility_identity_context=projected_facility_export_context(
                    projection,
                    (FacilityProjectionField.FACILITY_NAME,),
                ),
                facility_identity_conflicts=projected_facility_export_conflicts(
                    projection,
                    (FacilityProjectionField.FACILITY_NAME,),
                ),
            )
        )
    return projected


# ---------------------------------------------------------------------------
# CSV writers
# ---------------------------------------------------------------------------

_FACILITY_OVERVIEW_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "FacilityType",
    "Status",
    "City",
    "County",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "LoadedComplaintCount",
    "SubstantiatedOrEquivalentCount",
    "EarliestComplaintDate",
    "MostRecentComplaintDate",
    "LoadedDateRange",
    "SourceTraceabilityReadyCount",
    "MissingTraceabilityCount",
    "ComplaintDataLoadedStatus",
    "SourceSnapshotOrInput",
    "Limitations",
)

_SUBSTANTIATED_COMPLAINTS_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "ComplaintControlNumber",
    "ComplaintReceivedDate",
    "ReportDate",
    "FindingOrResolution",
    "SourceUrl",
    "RawHashOrArtifactReference",
    "ReviewerDetailPath",
    "Limitations",
)


def _write_facility_overview_csv(path: Path, facility_rows: list[_FacilityRow]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_FACILITY_OVERVIEW_FIELDS)
        writer.writeheader()
        for frow in facility_rows:
            dates = sorted(d for d in frow.complaint_dates if d)
            earliest = dates[0] if dates else UNKNOWN
            most_recent = dates[-1] if dates else UNKNOWN
            date_range = (
                f"{earliest} to {most_recent}"
                if dates
                else UNKNOWN
            )
            loaded_status = (
                "Complaints loaded"
                if frow.loaded_complaint_count > 0
                else "No complaints loaded"
            )
            writer.writerow(
                {
                    "FacilityNumber": frow.facility_number,
                    "FacilityName": frow.facility_name,
                    "FacilityType": frow.facility_type,
                    "Status": frow.status,
                    "City": frow.city,
                    "County": frow.county,
                    "FacilityIdentityContext": frow.facility_identity_context,
                    "FacilityIdentityConflicts": frow.facility_identity_conflicts,
                    "LoadedComplaintCount": frow.loaded_complaint_count,
                    "SubstantiatedOrEquivalentCount": frow.substantiated_count,
                    "EarliestComplaintDate": earliest,
                    "MostRecentComplaintDate": most_recent,
                    "LoadedDateRange": date_range,
                    "SourceTraceabilityReadyCount": frow.source_traceability_ready,
                    "MissingTraceabilityCount": frow.missing_traceability,
                    "ComplaintDataLoadedStatus": loaded_status,
                    "SourceSnapshotOrInput": UNKNOWN,
                    "Limitations": _limitations_sentence(),
                }
            )


def _write_substantiated_complaints_csv(
    path: Path, substantiated_rows: list[_SubstantiatedRow]
) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_SUBSTANTIATED_COMPLAINTS_FIELDS)
        writer.writeheader()
        for row in substantiated_rows:
            writer.writerow(
                {
                    "FacilityNumber": row.facility_number,
                    "FacilityName": row.facility_name,
                    "FacilityIdentityContext": row.facility_identity_context,
                    "FacilityIdentityConflicts": row.facility_identity_conflicts,
                    "ComplaintControlNumber": row.complaint_control_number,
                    "ComplaintReceivedDate": row.complaint_received_date,
                    "ReportDate": row.report_date,
                    "FindingOrResolution": row.finding_or_resolution,
                    "SourceUrl": row.source_url,
                    "RawHashOrArtifactReference": row.raw_hash_or_artifact_reference,
                    "ReviewerDetailPath": row.reviewer_detail_path,
                    "Limitations": row.limitations,
                }
            )


_COMPLAINT_RECORDS_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "FacilityType",
    "Status",
    "City",
    "County",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "ComplaintControlNumber",
    "ComplaintReceivedDate",
    "ReportDate",
    "FindingOrResolution",
    "FindingGroup",
    "ComplaintType",
    "AllegationCategory",
    "KeywordReviewCues",
    "SourceUrl",
    "RawHashOrArtifactReference",
    "ReviewerDetailPath",
    "Limitations",
)

_FACILITY_REVIEW_CUES_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "Status",
    "County",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "LoadedComplaintCount",
    "SubstantiatedOrEquivalentCount",
    "LoadedDateRange",
    "SourceTraceabilityReadyCount",
    "MissingTraceabilityCount",
    "ReviewCues",
    "SuggestedNextStep",
    "Limitations",
)


def _write_complaint_records_csv(
    path: Path, complaint_record_rows: list[_ComplaintRecordRow]
) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_COMPLAINT_RECORDS_FIELDS)
        writer.writeheader()
        for row in complaint_record_rows:
            writer.writerow(
                {
                    "FacilityNumber": row.facility_number,
                    "FacilityName": row.facility_name,
                    "FacilityType": row.facility_type,
                    "Status": row.status,
                    "City": row.city,
                    "County": row.county,
                    "FacilityIdentityContext": row.facility_identity_context,
                    "FacilityIdentityConflicts": row.facility_identity_conflicts,
                    "ComplaintControlNumber": row.complaint_control_number,
                    "ComplaintReceivedDate": row.complaint_received_date,
                    "ReportDate": row.report_date,
                    "FindingOrResolution": row.finding_or_resolution,
                    "FindingGroup": row.finding_group,
                    "ComplaintType": row.complaint_type,
                    "AllegationCategory": row.allegation_category,
                    "KeywordReviewCues": row.keyword_review_cues,
                    "SourceUrl": row.source_url,
                    "RawHashOrArtifactReference": row.raw_hash_or_artifact_reference,
                    "ReviewerDetailPath": row.reviewer_detail_path,
                    "Limitations": row.limitations,
                }
            )


# ---------------------------------------------------------------------------
# README
# ---------------------------------------------------------------------------

def _readme_text() -> str:
    return """\
# Stakeholder Facility Overview Extract

This extract is a review aid derived from locally loaded public CCLD complaint
records. It is not a certified report, legal finding, or source-completeness
proof.

## What this extract is

A summary of facility-level complaint counts and key fields drawn from records
that were loaded into the local SQLite database from publicly available CCLD
complaint report data.

## What this extract is not

- It does not make legal conclusions.
- It does not make facility-wide conclusions about a facility's conduct, safety,
  or compliance.
- It does not claim to be source-complete. Zero or low counts do not prove that
  no complaints exist; they reflect only what was loaded.
- It does not independently verify any finding or allegation. Finding/resolution
  values are source-derived values extracted from publicly available reports.
- It does not include verified severity determinations, risk scores, or rankings.
- It does not include raw narrative allegation text.

## Source of record

The public CCLD portal (ccld.dss.ca.gov) remains the source of record.
Verify important details against the public source before citing this extract.

## Counts and coverage

Counts are based only on currently loaded records. A facility may have additional
complaint records in the public source that were not retrieved or loaded.

## Facility identity

Public Facility ID and facility fields use the shared governed projection.
FacilityIdentityContext distinguishes current facility-reference values from
complaint-time observations. FacilityIdentityConflicts keeps differing governed
observations explicit. An unresolved numeric type such as 733 is shown only as
a source code whose descriptive label has not been verified. Internal facility
record identities are never substituted for the public Facility ID.

## Finding/resolution values

Finding/resolution values such as "Substantiated", "Founded", or "Sustained" are
extracted from publicly available CCLD complaint investigation reports and are
not independently verified by RecordsTracker. Do not restate them as independently
verified findings.

## Limitations column

Each row in facility-overview.csv, substantiated-complaints.csv, and
complaint-records.csv includes a Limitations column repeating this scope note
for use when rows are shared or excerpted outside this package.

## complaint-records.csv

This file contains one row per loaded complaint record for all facilities in
this extract, regardless of finding/resolution status.

- **FindingGroup** classifies each record into SubstantiatedOrEquivalent,
  NotSubstantiatedOrEquivalent, or UnknownOrMissing based on the source-derived
  finding value. This is not an independent legal determination.
- **ComplaintType** is a source-derived document type field when available;
  otherwise "not available".
- **AllegationCategory** is a source-derived category label when extracted;
  otherwise "not available". Raw narrative allegation text is not included.
- **KeywordReviewCues** is a deterministic keyword-based review-cue label
  derived from source-extracted non-narrative fields (finding, allegation
  category). A match signals a possible serious allegation topic as a review
  aid only. It is not a severity score, not a risk score, not a verified
  finding, and not a legal classification. A non-match does not confirm the
  absence of serious topics.
- Counts reflect only loaded records. Additional complaint records may exist
  in the public source that were not retrieved or loaded.
"""


# ---------------------------------------------------------------------------
# Manifest
# ---------------------------------------------------------------------------

def _build_manifest(
    *,
    generated_at: str,
    input_source: str,
    facility_row_count: int,
    substantiated_complaint_row_count: int,
    complaint_record_row_count: int,
    git_commit: str,
    output_files: list[str],
    limitations: str,
    facility_reference_csv: str,
    facility_reference_row_count: int,
    facility_reference_matched_count: int,
    only_facility_reference_rows: bool,
) -> dict[str, Any]:
    return {
        "generated_at": generated_at,
        "script_name": "export-stakeholder-facility-overview.ps1",
        "input_source": input_source,
        "output_files": output_files,
        "facility_row_count": facility_row_count,
        "substantiated_complaint_row_count": substantiated_complaint_row_count,
        "complaint_record_row_count": complaint_record_row_count,
        "git_commit": git_commit,
        "facility_reference_csv": facility_reference_csv,
        "facility_reference_row_count": facility_reference_row_count,
        "facility_reference_matched_count": facility_reference_matched_count,
        "only_facility_reference_rows": only_facility_reference_rows,
        "facility_identity_contract": (
            "Public Facility ID and facility presentation use the shared governed "
            "projection; current-reference and complaint-time context and conflicts "
            "remain explicit."
        ),
        "limitations": limitations,
    }


def _limitations_sentence() -> str:
    return (
        "Review aid over loaded source-derived records only. "
        "Does not make legal conclusions, facility-wide conclusions, "
        "source-completeness claims, verified severity claims, or abuse/neglect findings. "
        "Public CCLD portal remains the source of record."
    )


# ---------------------------------------------------------------------------
# Excel workbook writer
# ---------------------------------------------------------------------------

_XLSX_MAX_COL_WIDTH = 60

# ---------------------------------------------------------------------------
# Workbook presentation style constants
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_HEADER_BORDER_BOTTOM = Border(bottom=Side(style="thin"))
_TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
# Tab colours for the seven generated worksheets
_TAB_COLOR_README = "1F4E79"
_TAB_COLOR_SUMMARY = "7030A0"
_TAB_COLOR_FACILITY_REVIEW_CUES = "4472C4"
_TAB_COLOR_FACILITY_OVERVIEW = "2E75B6"
_TAB_COLOR_SUBSTANTIATED = "C55A11"
_TAB_COLOR_COMPLAINT_RECORDS = "538135"
_TAB_COLOR_MANIFEST = "767171"
# Column names that receive wrap_text alignment in data worksheets.
_WRAP_COLS = frozenset(
    {
        "Limitations",
        "ReviewCues",
        "SuggestedNextStep",
        "FacilityIdentityContext",
        "FacilityIdentityConflicts",
    }
)
# Display headers for XLSX worksheets (stakeholder-friendly spacing/capitalisation).
_XLSX_DISPLAY_HEADERS: dict[str, str] = {
    "FacilityNumber": "Facility Number",
    "FacilityName": "Facility Name",
    "FacilityType": "Facility Type",
    "FacilityIdentityContext": "Facility Identity Context",
    "FacilityIdentityConflicts": "Facility Identity Conflicts",
    "LoadedComplaintCount": "Loaded Complaint Count",
    "SubstantiatedOrEquivalentCount": "Substantiated Count",
    "EarliestComplaintDate": "Earliest Complaint Date",
    "MostRecentComplaintDate": "Most Recent Complaint Date",
    "LoadedDateRange": "Loaded Date Range",
    "SourceTraceabilityReadyCount": "Source Traceability Ready Count",
    "MissingTraceabilityCount": "Missing Traceability Count",
    "ComplaintDataLoadedStatus": "Complaint Data Status",
    "ComplaintControlNumber": "Complaint Control Number",
    "ComplaintReceivedDate": "Complaint Received Date",
    "ReportDate": "Report Date",
    "FindingOrResolution": "Finding / Resolution",
    "FindingGroup": "Finding Group",
    "ComplaintType": "Complaint Type",
    "AllegationCategory": "Allegation Category",
    "KeywordReviewCues": "Keyword Review Cues",
    "SourceUrl": "Source URL",
    "ReviewCues": "Review Cues",
    "SuggestedNextStep": "Suggested Next Step",
}
# XLSX display field tuples — subset of CSV fields for the workbook.
# Columns removed: Limitations from all sheets; RawHashOrArtifactReference and
# ReviewerDetailPath from complaint sheets; SourceSnapshotOrInput (always UNKNOWN).
_FACILITY_REVIEW_CUES_XLSX_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "Status",
    "County",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "LoadedComplaintCount",
    "SubstantiatedOrEquivalentCount",
    "LoadedDateRange",
    "SourceTraceabilityReadyCount",
    "MissingTraceabilityCount",
    "ReviewCues",
    "SuggestedNextStep",
)
_FACILITY_OVERVIEW_XLSX_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "FacilityType",
    "Status",
    "City",
    "County",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "LoadedComplaintCount",
    "SubstantiatedOrEquivalentCount",
    "EarliestComplaintDate",
    "MostRecentComplaintDate",
    "LoadedDateRange",
    "SourceTraceabilityReadyCount",
    "MissingTraceabilityCount",
    "ComplaintDataLoadedStatus",
)
_SUBSTANTIATED_COMPLAINTS_XLSX_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "ComplaintControlNumber",
    "ComplaintReceivedDate",
    "ReportDate",
    "FindingOrResolution",
    "SourceUrl",
)
_COMPLAINT_RECORDS_XLSX_FIELDS = (
    "FacilityNumber",
    "FacilityName",
    "FacilityType",
    "Status",
    "City",
    "County",
    "FacilityIdentityContext",
    "FacilityIdentityConflicts",
    "ComplaintControlNumber",
    "ComplaintReceivedDate",
    "ReportDate",
    "FindingOrResolution",
    "FindingGroup",
    "ComplaintType",
    "AllegationCategory",
    "KeywordReviewCues",
    "SourceUrl",
)
# Display labels for the XLSX Manifest worksheet (maps manifest.json keys to
# stakeholder-readable labels).  The external manifest.json is unchanged.
_MANIFEST_DISPLAY_LABELS: dict[str, str] = {
    "generated_at": "Generated",
    "script_name": "Script Name",
    "input_source": "Input Source",
    "output_files": "Output Files",
    "facility_row_count": "Facility Row Count",
    "substantiated_complaint_row_count": "Substantiated Complaint Row Count",
    "complaint_record_row_count": "Complaint Record Row Count",
    "git_commit": "Git Commit",
    "facility_reference_csv": "Facility Reference CSV",
    "facility_reference_row_count": "Facility Reference Row Count",
    "facility_reference_matched_count": "Facility Reference Matched Count",
    "only_facility_reference_rows": "Reference-Only Filter Applied",
    "facility_identity_contract": "Facility Identity Contract",
    "limitations": "Limitations",
}


def _facility_overview_row_dicts(
    facility_rows: list[_FacilityRow],
) -> list[dict[str, Any]]:
    """Build row dicts for the facility-overview worksheet."""
    lim = _limitations_sentence()
    result: list[dict[str, Any]] = []
    for frow in facility_rows:
        dates = sorted(d for d in frow.complaint_dates if d)
        earliest = dates[0] if dates else UNKNOWN
        most_recent = dates[-1] if dates else UNKNOWN
        date_range = f"{earliest} to {most_recent}" if dates else UNKNOWN
        loaded_status = (
            "Complaints loaded" if frow.loaded_complaint_count > 0 else "No complaints loaded"
        )
        result.append(
            {
                "FacilityNumber": frow.facility_number,
                "FacilityName": frow.facility_name,
                "FacilityType": frow.facility_type,
                "Status": frow.status,
                "City": frow.city,
                "County": frow.county,
                "FacilityIdentityContext": frow.facility_identity_context,
                "FacilityIdentityConflicts": frow.facility_identity_conflicts,
                "LoadedComplaintCount": frow.loaded_complaint_count,
                "SubstantiatedOrEquivalentCount": frow.substantiated_count,
                "EarliestComplaintDate": earliest,
                "MostRecentComplaintDate": most_recent,
                "LoadedDateRange": date_range,
                "SourceTraceabilityReadyCount": frow.source_traceability_ready,
                "MissingTraceabilityCount": frow.missing_traceability,
                "ComplaintDataLoadedStatus": loaded_status,
                "SourceSnapshotOrInput": UNKNOWN,
                "Limitations": lim,
            }
        )
    return result


def _substantiated_row_dicts(
    substantiated_rows: list[_SubstantiatedRow],
) -> list[dict[str, Any]]:
    """Build row dicts for the substantiated-complaints worksheet."""
    return [
        {
            "FacilityNumber": row.facility_number,
            "FacilityName": row.facility_name,
            "FacilityIdentityContext": row.facility_identity_context,
            "FacilityIdentityConflicts": row.facility_identity_conflicts,
            "ComplaintControlNumber": row.complaint_control_number,
            "ComplaintReceivedDate": row.complaint_received_date,
            "ReportDate": row.report_date,
            "FindingOrResolution": row.finding_or_resolution,
            "SourceUrl": row.source_url,
            "RawHashOrArtifactReference": row.raw_hash_or_artifact_reference,
            "ReviewerDetailPath": row.reviewer_detail_path,
            "Limitations": row.limitations,
        }
        for row in substantiated_rows
    ]


def _complaint_record_row_dicts(
    complaint_record_rows: list[_ComplaintRecordRow],
) -> list[dict[str, Any]]:
    """Build row dicts for the complaint-records worksheet."""
    return [
        {
            "FacilityNumber": row.facility_number,
            "FacilityName": row.facility_name,
            "FacilityType": row.facility_type,
            "Status": row.status,
            "City": row.city,
            "County": row.county,
            "FacilityIdentityContext": row.facility_identity_context,
            "FacilityIdentityConflicts": row.facility_identity_conflicts,
            "ComplaintControlNumber": row.complaint_control_number,
            "ComplaintReceivedDate": row.complaint_received_date,
            "ReportDate": row.report_date,
            "FindingOrResolution": row.finding_or_resolution,
            "FindingGroup": row.finding_group,
            "ComplaintType": row.complaint_type,
            "AllegationCategory": row.allegation_category,
            "KeywordReviewCues": row.keyword_review_cues,
            "SourceUrl": row.source_url,
            "RawHashOrArtifactReference": row.raw_hash_or_artifact_reference,
            "ReviewerDetailPath": row.reviewer_detail_path,
            "Limitations": row.limitations,
        }
        for row in complaint_record_rows
    ]


def _derive_facility_review_cues(frow: _FacilityRow) -> tuple[str, str]:
    """Return (review_cues_str, suggested_next_step) for a facility row.

    Review cues are deterministic labels derived from loaded record counts only.
    They are review aids, not risk scores, rankings, or legal conclusions.
    """
    cues: list[str] = []
    if frow.substantiated_count > 0:
        cues.append("Substantiated complaint records loaded")
    if frow.loaded_complaint_count > 1:
        cues.append("Multiple loaded complaint records")
    if frow.loaded_complaint_count == 0:
        cues.append("No complaint records loaded in this extract")
    if frow.missing_traceability > 0:
        cues.append("Missing source traceability in loaded rows")
    if (
        frow.status.strip().casefold().startswith("closed")
        and frow.loaded_complaint_count > 0
    ):
        cues.append("Closed facility with loaded complaint records")
    review_cues = "; ".join(cues) if cues else "No review cues for loaded data"
    if frow.loaded_complaint_count == 0:
        suggested = (
            "Confirm whether no loaded records reflects extract scope "
            "before drawing any conclusion."
        )
    else:
        suggested = (
            "Review Facility Overview and Complaint Records rows; verify "
            "important details against the public CCLD portal."
        )
    return review_cues, suggested


def _facility_review_cues_row_dicts(
    facility_rows: list[_FacilityRow],
) -> list[dict[str, Any]]:
    """Build row dicts for the facility-review-cues worksheet."""
    lim = _limitations_sentence()
    result: list[dict[str, Any]] = []
    for frow in facility_rows:
        dates = sorted(d for d in frow.complaint_dates if d)
        earliest = dates[0] if dates else UNKNOWN
        most_recent = dates[-1] if dates else UNKNOWN
        date_range = f"{earliest} to {most_recent}" if dates else UNKNOWN
        review_cues, suggested = _derive_facility_review_cues(frow)
        result.append(
            {
                "FacilityNumber": frow.facility_number,
                "FacilityName": frow.facility_name,
                "Status": frow.status,
                "County": frow.county,
                "FacilityIdentityContext": frow.facility_identity_context,
                "FacilityIdentityConflicts": frow.facility_identity_conflicts,
                "LoadedComplaintCount": frow.loaded_complaint_count,
                "SubstantiatedOrEquivalentCount": frow.substantiated_count,
                "LoadedDateRange": date_range,
                "SourceTraceabilityReadyCount": frow.source_traceability_ready,
                "MissingTraceabilityCount": frow.missing_traceability,
                "ReviewCues": review_cues,
                "SuggestedNextStep": suggested,
                "Limitations": lim,
            }
        )
    return result


def _finding_group_display(fg: str) -> str:
    """Map internal FindingGroup labels to stakeholder-readable display values."""
    return {
        "SubstantiatedOrEquivalent": "Substantiated",
        "NotSubstantiatedOrEquivalent": "Unsubstantiated",
    }.get(fg, fg)


def _xlsx_complaint_record_xlsx_row_dicts(
    complaint_record_rows: list[_ComplaintRecordRow],
) -> list[dict[str, Any]]:
    """Build XLSX row dicts for complaint-records with stakeholder-readable FindingGroup."""
    rows = _complaint_record_row_dicts(complaint_record_rows)
    for row in rows:
        row["FindingGroup"] = _finding_group_display(row.get("FindingGroup", ""))
    return rows


def _format_generated_at(timestamp: str) -> str:
    """Convert '20260621T215952Z' to a human-readable form like 'June 21, 2026, 9:59 PM UTC'."""
    try:
        dt = datetime.strptime(timestamp, "%Y%m%dT%H%M%SZ").replace(tzinfo=UTC)
        hour = dt.strftime("%I").lstrip("0") or "12"
        minute = dt.strftime("%M")
        ampm = dt.strftime("%p")
        return f"{dt.strftime('%B')} {dt.day}, {dt.year}, {hour}:{minute} {ampm} UTC"
    except ValueError:
        return timestamp


def _xlsx_autosize_columns(ws: Any, col_letter_fn: Any) -> None:
    """Auto-size worksheet columns, capped at _XLSX_MAX_COL_WIDTH characters."""
    for col_cells in ws.columns:
        col_letter = col_letter_fn(col_cells[0].column)
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=0,
        )
        ws.column_dimensions[col_letter].width = max(
            min(max_len + 2, _XLSX_MAX_COL_WIDTH), 8
        )


def _xlsx_data_sheet(
    ws: Any,
    *,
    fields: tuple[str, ...],
    rows: list[dict[str, Any]],
    bold_font: Any,
    col_letter_fn: Any,
    display_headers: dict[str, str] | None = None,
    freeze_col: str = "A",
) -> None:
    """Write bold headers and data rows to a worksheet.

    Freezes the header row (and optionally identifying left columns via freeze_col),
    applies auto-filter, and auto-sizes columns.
    Applies top-left alignment to all cells; wrap_text for columns in _WRAP_COLS.
    """
    _dh = display_headers or {}
    _top_left = Alignment(horizontal="left", vertical="top")
    for col_idx, field_name in enumerate(fields, 1):
        display = _dh.get(field_name, field_name)
        cell = ws.cell(row=1, column=col_idx, value=display)
        cell.font = bold_font
        cell.fill = _HEADER_FILL
        cell.border = _HEADER_BORDER_BOTTOM
        cell.alignment = _top_left
    for row_dict in rows:
        ws.append([row_dict.get(f, "") for f in fields])
    ws.freeze_panes = f"{freeze_col}2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    # Build column maps for per-cell formatting.
    source_url_col: int | None = None
    wrap_col_indices: set[int] = set()
    for _col_idx, _fname in enumerate(fields, 1):
        if _fname == "SourceUrl":
            source_url_col = _col_idx
        if _fname in _WRAP_COLS:
            wrap_col_indices.add(_col_idx)
    # Apply top-left alignment to all data cells; wrap long-text columns.
    for _row_idx in range(2, ws.max_row + 1):
        for _col_idx in range(1, len(fields) + 1):
            _cell = ws.cell(row=_row_idx, column=_col_idx)
            _wrap = _col_idx in wrap_col_indices
            _cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=_wrap)
        if source_url_col is not None:
            _src_cell = ws.cell(row=_row_idx, column=source_url_col)
            _val = str(_src_cell.value or "")
            if _val.startswith("http"):
                _src_cell.hyperlink = _val
    _xlsx_autosize_columns(ws, col_letter_fn)


def _xlsx_populate_readme_sheet(
    ws: Any,
    *,
    manifest: dict[str, Any],
    bold_font: Any,
) -> None:
    """Populate the README worksheet with scope, limitations, and key details."""
    _narrative_rows: list[int] = []

    def _row(col_a: str, col_b: str = "") -> None:
        ws.append([col_a, col_b])

    def _narrative(text: str) -> None:
        """Append a prose-only row; tracked for A:B merge after all rows are written."""
        ws.append([text, ""])
        _narrative_rows.append(ws.max_row)

    def _section(title: str) -> None:
        ws.append([""])
        cell = ws.cell(row=ws.max_row, column=1, value=title)
        cell.font = bold_font
        cell.fill = _HEADER_FILL
        ws.cell(row=ws.max_row, column=2).fill = _HEADER_FILL

    _row("Stakeholder Facility Overview — Review Aid")
    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.cell(row=1, column=1).fill = _TITLE_FILL
    ws.cell(row=1, column=2).fill = _TITLE_FILL

    _section("PURPOSE")
    _narrative(
        "This workbook is a review aid derived from locally loaded public CCLD "
        "complaint records. It is not a certified report, legal finding, or "
        "source-completeness proof."
    )

    _section("KEY DETAILS")
    _row("Generated", _format_generated_at(str(manifest.get("generated_at", ""))))
    _row("Git commit", str(manifest.get("git_commit", "")))
    _row("Facilities", str(manifest.get("facility_row_count", 0)))
    _row(
        "Substantiated complaints",
        str(manifest.get("substantiated_complaint_row_count", 0)),
    )
    _row("All loaded complaint records", str(manifest.get("complaint_record_row_count", 0)))
    _row("Facility reference CSV", str(manifest.get("facility_reference_csv", "none")))
    _row("Facility reference rows", str(manifest.get("facility_reference_row_count", 0)))
    _row(
        "Reference rows matched to loaded complaints",
        str(manifest.get("facility_reference_matched_count", 0)),
    )
    _row(
        "Reference-only filter applied",
        str(manifest.get("only_facility_reference_rows", False)),
    )

    _section("HOW TO USE THIS WORKBOOK")
    _narrative(
        "Start with the Facility Overview tab for a per-facility summary. "
        "Use Substantiated Complaints for individual records where the source-derived "
        "finding is substantiated. "
        "Use Complaint Records for all loaded complaint records regardless of finding status. "
        "Check the Manifest tab for generation metadata. "
        "Verify important details against the public CCLD portal before citing this extract."
    )

    _section("TABS INCLUDED")
    _row("README", "This tab — scope, limitations, how to use this workbook.")
    _row(
        "Summary",
        "Source-derived workbook metrics: key counts, facility status breakdown, "
        "finding group breakdown, and keyword review cue breakdown.",
    )
    _row(
        "Facility Review Cues",
        "One row per facility with source-derived review cue labels and a "
        "suggested next step. Review aids only; does not rank or score facilities.",
    )
    _row(
        "Facility Overview",
        "Per-facility complaint counts, substantiated counts, "
        "date ranges, and source-traceability counts.",
    )
    _row(
        "Substantiated Complaints",
        "Individual complaint records where the source-derived finding "
        "indicates substantiated.",
    )
    _row(
        "Complaint Records",
        "All loaded complaint records regardless of finding status. "
        "Includes Finding Group, Complaint Type, Allegation Category, Keyword Review Cues.",
    )
    _row("Manifest", "Generation metadata, row counts, and output file list.")

    _section("COUNTS AND COVERAGE")
    _narrative(
        "Counts are based only on records currently loaded in the local database. "
        "A facility may have additional complaint records in the public source that "
        "were not retrieved or loaded. Zero or low counts do not prove that no "
        "complaints exist; they reflect only what was loaded."
    )

    _section("SOURCE OF RECORD")
    _narrative(
        "The public CCLD portal (ccld.dss.ca.gov) remains the source of record. "
        "Verify important details against the public source before citing this extract."
    )

    _section("IMPORTANT LIMITATIONS")
    _narrative(
        "This extract does not make legal conclusions, facility-wide conclusions, "
        "source-completeness claims, verified severity claims, or abuse/neglect findings. "
        "Finding/resolution values are source-derived and not independently verified by "
        "RecordsTracker. Raw narrative allegation text is intentionally excluded."
    )
    _narrative(
        "Keyword Review Cues is a deterministic keyword-based review-cue label derived from "
        "source-extracted non-narrative fields (finding, allegation category). A match "
        "signals a possible serious allegation topic as a review aid only. It is not a "
        "severity score, not a risk score, not a verified finding, and not a legal "
        "classification. A non-match does not confirm the absence of serious topics."
    )
    _row("")
    _narrative(str(manifest.get("limitations", "")))

    # Merge narrative rows across A:B so prose spans the full row width.
    for _rn in _narrative_rows:
        ws.merge_cells(f"A{_rn}:B{_rn}")

    # Wrap all cell text in the README so long prose is readable.
    for _ws_row in ws.iter_rows():
        for _cell in _ws_row:
            _cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 80


def _xlsx_populate_summary_sheet(
    ws: Any,
    *,
    manifest: dict[str, Any],
    facility_rows: list[_FacilityRow],
    complaint_record_rows: list[_ComplaintRecordRow],
    bold_font: Any,
) -> None:
    """Populate the Summary worksheet with source-derived workbook metrics.

    All counts are derived from records loaded locally.  Does not make legal
    conclusions, facility-wide conclusions, source-completeness claims, risk
    scores, severity scores, or rankings.
    """
    _narrative_rows: list[int] = []

    def _kv(label: str, value: str = "") -> None:
        ws.append([label, value])

    def _narrative(text: str) -> None:
        """Append a prose-only row; tracked for A:B merge after all rows are written."""
        ws.append([text, ""])
        _narrative_rows.append(ws.max_row)

    def _section(title: str) -> None:
        ws.append([""])
        cell = ws.cell(row=ws.max_row, column=1, value=title)
        cell.font = bold_font
        cell.fill = _HEADER_FILL
        ws.cell(row=ws.max_row, column=2).fill = _HEADER_FILL

    # Title
    ws.append(["Summary — Source-Derived Workbook Metrics", ""])
    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.cell(row=1, column=1).fill = _TITLE_FILL
    ws.cell(row=1, column=2).fill = _TITLE_FILL

    _section("HOW TO USE THIS SUMMARY")
    _narrative(
        "Counts and groups are based only on records currently loaded in the local "
        "database. A facility may have additional complaint records in the public "
        "source that were not retrieved or loaded. Zero or low counts do not prove "
        "that no complaints exist. This summary does not make legal conclusions, "
        "facility-wide conclusions, source-completeness claims, verified severity "
        "claims, or abuse/neglect findings. Verify important details against the "
        "public CCLD portal before citing this extract."
    )

    _section("KEY METRICS (source-derived count)")
    _kv("Generated", _format_generated_at(str(manifest.get("generated_at", ""))))
    _kv("Facilities", str(manifest.get("facility_row_count", 0)))
    _kv("Loaded complaint records", str(manifest.get("complaint_record_row_count", 0)))
    _kv(
        "Substantiated complaints",
        str(manifest.get("substantiated_complaint_row_count", 0)),
    )
    _kv("Facility reference rows", str(manifest.get("facility_reference_row_count", 0)))
    _kv(
        "Facility reference rows matched to loaded complaints",
        str(manifest.get("facility_reference_matched_count", 0)),
    )

    _section("FACILITY STATUS COUNTS (source-derived, Facility Overview tab)")
    status_counts: dict[str, int] = {}
    for _frow in facility_rows:
        status_counts[_frow.status] = status_counts.get(_frow.status, 0) + 1
    if status_counts:
        for _status_val, _count in sorted(status_counts.items()):
            _kv(f"  {_status_val}", str(_count))
    else:
        _kv("  (no facilities loaded)")

    _section("FINDING GROUP COUNTS (source-derived, Complaint Records tab)")
    fg_counts: dict[str, int] = {}
    for _cr in complaint_record_rows:
        fg_counts[_cr.finding_group] = fg_counts.get(_cr.finding_group, 0) + 1
    if fg_counts:
        for _fg_val, _fg_count in sorted(fg_counts.items()):
            _kv(f"  {_finding_group_display(_fg_val)}", str(_fg_count))
    else:
        _kv("  (no complaint records loaded)")

    _section("KEYWORD REVIEW CUE COUNTS (source-derived review aid, Complaint Records tab)")
    cue_counts: dict[str, int] = {}
    for _cr in complaint_record_rows:
        cue_label = _cr.keyword_review_cues.strip() if _cr.keyword_review_cues else ""
        if not cue_label:
            cue_label = "(no review cue match)"
        cue_counts[cue_label] = cue_counts.get(cue_label, 0) + 1
    _has_cue_match = any(
        label not in (UNKNOWN, "(no review cue match)")
        for label in cue_counts
    )
    if not complaint_record_rows:
        _kv("  (no complaint records loaded)")
    elif not _has_cue_match:
        _narrative(
            "Keyword review cues were not available in this extract. "
            "No loaded records matched the deterministic keyword list."
        )
    else:
        for _cue_label, _cue_count in sorted(cue_counts.items()):
            _kv(f"  {_cue_label}", str(_cue_count))
    _narrative(
        "Note: a keyword review cue match signals a possible serious allegation "
        "topic as a review aid only. It is not a severity score, not a risk score, "
        "not a verified finding, and not a legal classification."
    )

    _section("SOURCE OF RECORD")
    _narrative(
        "The public CCLD portal (ccld.dss.ca.gov) remains the source of record. "
        "Verify important details against the public source before citing this extract."
    )

    # Merge narrative rows across A:B so prose spans the full row width.
    for _rn in _narrative_rows:
        ws.merge_cells(f"A{_rn}:B{_rn}")

    # Wrap all cell text so long prose is readable.
    for _ws_row in ws.iter_rows():
        for _cell in _ws_row:
            _cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 20


def _write_xlsx_workbook(
    xlsx_path: Path,
    *,
    facility_rows: list[_FacilityRow],
    substantiated_rows: list[_SubstantiatedRow],
    complaint_record_rows: list[_ComplaintRecordRow],
    manifest: dict[str, Any],
) -> None:
    """Write the stakeholder Excel workbook.

    Sheet order: README, Summary, Facility Review Cues, Facility Overview,
    Substantiated Complaints, Complaint Records, Manifest.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)  # remove the default "Sheet"
    bold = Font(bold=True)

    ws_readme = wb.create_sheet("README")
    ws_readme.sheet_properties.tabColor = _TAB_COLOR_README
    _xlsx_populate_readme_sheet(ws_readme, manifest=manifest, bold_font=bold)

    ws_summary = wb.create_sheet("Summary")
    ws_summary.sheet_properties.tabColor = _TAB_COLOR_SUMMARY
    _xlsx_populate_summary_sheet(
        ws_summary,
        manifest=manifest,
        facility_rows=facility_rows,
        complaint_record_rows=complaint_record_rows,
        bold_font=bold,
    )

    ws_frc = wb.create_sheet("Facility Review Cues")
    ws_frc.sheet_properties.tabColor = _TAB_COLOR_FACILITY_REVIEW_CUES
    _xlsx_data_sheet(
        ws_frc,
        fields=_FACILITY_REVIEW_CUES_XLSX_FIELDS,
        rows=_facility_review_cues_row_dicts(facility_rows),
        bold_font=bold,
        col_letter_fn=get_column_letter,
        display_headers=_XLSX_DISPLAY_HEADERS,
        freeze_col="C",
    )

    ws_fo = wb.create_sheet("Facility Overview")
    ws_fo.sheet_properties.tabColor = _TAB_COLOR_FACILITY_OVERVIEW
    _xlsx_data_sheet(
        ws_fo,
        fields=_FACILITY_OVERVIEW_XLSX_FIELDS,
        rows=_facility_overview_row_dicts(facility_rows),
        bold_font=bold,
        col_letter_fn=get_column_letter,
        display_headers=_XLSX_DISPLAY_HEADERS,
        freeze_col="C",
    )

    ws_sub = wb.create_sheet("Substantiated Complaints")
    ws_sub.sheet_properties.tabColor = _TAB_COLOR_SUBSTANTIATED
    _xlsx_data_sheet(
        ws_sub,
        fields=_SUBSTANTIATED_COMPLAINTS_XLSX_FIELDS,
        rows=_substantiated_row_dicts(substantiated_rows),
        bold_font=bold,
        col_letter_fn=get_column_letter,
        display_headers=_XLSX_DISPLAY_HEADERS,
        freeze_col="C",
    )

    ws_cr = wb.create_sheet("Complaint Records")
    ws_cr.sheet_properties.tabColor = _TAB_COLOR_COMPLAINT_RECORDS
    _xlsx_data_sheet(
        ws_cr,
        fields=_COMPLAINT_RECORDS_XLSX_FIELDS,
        rows=_xlsx_complaint_record_xlsx_row_dicts(complaint_record_rows),
        bold_font=bold,
        col_letter_fn=get_column_letter,
        display_headers=_XLSX_DISPLAY_HEADERS,
        freeze_col="C",
    )

    ws_manifest = wb.create_sheet("Manifest")
    ws_manifest.sheet_properties.tabColor = _TAB_COLOR_MANIFEST
    ws_manifest.append(["Field", "Value"])
    ws_manifest.cell(row=1, column=1).font = bold
    ws_manifest.cell(row=1, column=1).fill = _HEADER_FILL
    ws_manifest.cell(row=1, column=1).border = _HEADER_BORDER_BOTTOM
    ws_manifest.cell(row=1, column=2).font = bold
    ws_manifest.cell(row=1, column=2).fill = _HEADER_FILL
    ws_manifest.cell(row=1, column=2).border = _HEADER_BORDER_BOTTOM
    for key, value in manifest.items():
        display_key = _MANIFEST_DISPLAY_LABELS.get(key, key)
        if key == "generated_at":
            display_val = _format_generated_at(str(value or ""))
        elif isinstance(value, list):
            display_val = "; ".join(str(v) for v in value)
        else:
            display_val = str(value) if value is not None else ""
        ws_manifest.append([display_key, display_val])
    ws_manifest.freeze_panes = "A2"
    ws_manifest.auto_filter.ref = ws_manifest.dimensions
    ws_manifest.column_dimensions["A"].width = 40
    ws_manifest.column_dimensions["B"].width = 80
    # Apply top-left alignment to all Manifest cells.
    for _ws_row in ws_manifest.iter_rows():
        for _cell in _ws_row:
            _cell.alignment = Alignment(horizontal="left", vertical="top")

    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Git commit helper
# ---------------------------------------------------------------------------

def _get_git_commit() -> str:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            check=False,
            timeout=5,
        )
        commit = result.stdout.strip()
        return commit if commit else "unknown"
    except Exception:
        return "unknown"
